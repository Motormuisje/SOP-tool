"""
S&OP Planning Engine - Master workbook (per-site Excel editing medium).

The app master store stays the runtime source of truth; this module renders
it to a clean, per-site Excel workbook (one table per dataset, columns =
dataclass field names) and parses an edited copy back into a master dict.
The legacy MS_RECONC format is deliberately NOT reproduced here — this is
the app's own stable format, guarded by a round-trip test
(export -> parse == original master) and by validation-through-hydration on
import (same mechanism as the dataset PATCH route).

Month data stays out by design: purchase ACTUALS are not exported, and an
import preserves the store's existing actuals untouched (F2 principle:
master data and month data must not mix silently).

The _Meta sheet carries site + store version so an import can refuse a
workbook from another site and detect edits based on a stale export.
"""

import dataclasses
import math
import re
import typing
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from modules.master_data import FTE_DATASETS
from modules.models import (
    FTE_PARAM_DEFAULTS,
    Machine,
    MachineCombination,
    MachineCost,
    Material,
    RawMaterialCost,
    SafetyStockConfig,
    SalesPriceItem,
    ValuationParameters,
)

WORKBOOK_SCHEMA_VERSION = 1

SHEET_META = '_Meta'
SHEET_CONFIG = 'Config'
SHEET_FTE = 'FTE'
SHEET_MATERIALS = 'Materialen'
SHEET_MACHINES = 'Machines'
SHEET_SAFETY = 'Veiligheidsvoorraad'
SHEET_PURCHASE = 'Inkoop'
SHEET_PRICES = 'Verkoopprijzen'
SHEET_MATERIAL_COSTS = 'Grondstofkosten'
SHEET_MACHINE_COSTS = 'Machinekosten'
SHEET_VALUATION = 'Waardering'

# F2-CF workbench datasets. Sheet per dataset, same keyed-table shape as the
# rest. These sheets are OPTIONAL on import: a workbook exported by an older
# build has none of them, and refusing it would strand every existing copy.
SHEET_STAFFING = 'Bemensing'
SHEET_LABOR_RATES = 'Loonkosten'
SHEET_COMBINATIONS = 'Machinecombinaties'
SHEET_INDIRECT = 'Indirecte activiteiten'
SHEET_THROUGHPUT = 'Doorzet-overrides'
SHEET_BENCHMARK = 'Benchmark doorzet'

SHEET_CHANGEOVER = 'Omsteltijden'

FTE_DATASET_SHEETS = {
    'staffing_norms': SHEET_STAFFING,
    'labor_rates': SHEET_LABOR_RATES,
    'machine_combinations': SHEET_COMBINATIONS,
    'indirect_activities': SHEET_INDIRECT,
    'throughput_overrides': SHEET_THROUGHPUT,
    'benchmark_throughput': SHEET_BENCHMARK,
    'changeover_times': SHEET_CHANGEOVER,
}

_KEY_COLUMN = 'sleutel'
_AVAILABILITY_PREFIX = 'beschikbaarheid '
_FTE_PARAM_PREFIX = 'params.'


def _field_names(dc) -> List[str]:
    return [f.name for f in dataclasses.fields(dc)]


def _casters(dc) -> Dict[str, callable]:
    """Per-field cell->value casters derived from the dataclass type hints.

    Cells come back from Excel as str/float/bool/None; the master dict must
    carry the same shapes serialize_master produces. Enum-valued fields
    (product_type, shift_system) are stored as their string value and stay
    strings here — hydration converts them.
    """
    hints = typing.get_type_hints(dc)
    defaults = {}
    for f in dataclasses.fields(dc):
        if f.default is not dataclasses.MISSING:
            defaults[f.name] = f.default
        elif f.default_factory is not dataclasses.MISSING:  # type: ignore[misc]
            # list/dict fields (machine_codes, throughput_factor_by_machine)
            # carry their default through a factory; without this an empty
            # cell fell through to '' and hydration raised on the wrong type.
            defaults[f.name] = f.default_factory()  # type: ignore[misc]
    casters = {}
    for name, hint in hints.items():
        origin = typing.get_origin(hint)
        args = typing.get_args(hint)
        optional = origin is typing.Union and type(None) in args
        if optional:
            inner = next(a for a in args if a is not type(None))
        else:
            inner = hint
        casters[name] = _make_caster(inner, optional, defaults.get(name, dataclasses.MISSING))
    return casters


def _empty_value(inner, origin, optional, default):
    """What an empty cell means: the dataclass default, else the type's zero."""
    if default is not dataclasses.MISSING:
        # Copy: a shared mutable default would let two rows of the same sheet
        # alias one list.
        if isinstance(default, list):
            return list(default)
        if isinstance(default, dict):
            return dict(default)
        return default.value if hasattr(default, 'value') else default
    if optional:
        return None
    if inner is bool:
        return False
    if inner is int:
        return 0
    if inner is float:
        return 0.0
    if origin in (list, set, tuple):
        return []
    if origin is dict:
        return {}
    return ''


def _make_caster(inner, optional, default):
    origin = typing.get_origin(inner)
    args = typing.get_args(inner)

    def cast(value):
        if value is None or (isinstance(value, str) and not value.strip()):
            return _empty_value(inner, origin, optional, default)
        if inner in (bool, int, float, str):
            return _scalar(inner, value)
        if origin in (list, set, tuple):
            return [_scalar(args[0] if args else str, part)
                    for part in _split_cells(value)]
        if origin is dict:
            return _parse_pairs(value, args[1] if len(args) > 1 else str)
        return value  # enums-as-strings — passed through
    return cast


def _split_cells(value) -> List[str]:
    return [part.strip() for part in str(value).split(',') if part.strip()]


def _scalar(item_type, text):
    """Cell/fragment -> scalar. Deliberately as strict as the pre-F2-CF caster:
    a decimal comma stays an error, because inside a list/map cell the comma is
    the separator and '0,8' cannot mean both."""
    if item_type is float:
        return float(text)
    if item_type is int:
        return int(float(text))
    if item_type is bool:
        return _to_bool(text)
    return str(text).strip()


def _parse_pairs(value, value_type) -> dict:
    """'PBA01:0.8, PBA02:1' -> {'PBA01': 0.8, 'PBA02': 1.0}.

    A malformed fragment is a typo in a hand-edited sheet and must be rejected,
    not silently dropped: a dropped per-machine factor reads as 'no effect'.
    """
    out = {}
    for fragment in _split_cells(value):
        parts = fragment.split(':')
        if len(parts) != 2 or not parts[0].strip():
            raise MasterWorkbookError(
                f'"{fragment}" is geen "SLEUTEL:waarde"-paar.')
        try:
            out[parts[0].strip()] = _scalar(value_type, parts[1])
        except ValueError as exc:
            raise MasterWorkbookError(
                f'Waarde bij "{parts[0].strip()}" is geen getal '
                f'("{parts[1].strip()}").') from exc
    return out


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ('1', 'true', 'waar', 'ja', 'x', 'yes')


def _row_dict(headers, row) -> dict:
    return {h: row[i] if i < len(row) else None for i, h in enumerate(headers) if h}


def _is_empty_row(values) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


# ------------------------------------------------------------------ export


def export_master_workbook(master: dict, path, site: str,
                           store_version, exported_at: Optional[str] = None) -> None:
    """Render the master dict to the per-site workbook at `path` (atomic:
    written to a temp name first, then replaced)."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_kv(wb, SHEET_META, [
        ('schema_version', WORKBOOK_SCHEMA_VERSION),
        ('site', site),
        ('store_version', store_version),
        ('exported_at', exported_at or datetime.now().isoformat(timespec='seconds')),
        ('let_op', 'Gegenereerd door de app. Bewerk de gegevensbladen; laat _Meta ongemoeid.'),
    ])

    cfg = master.get('config') or {}
    _write_kv(wb, SHEET_CONFIG, [
        ('initial_date', cfg.get('initial_date') or ''),
        ('forecast_months', cfg.get('forecast_months')),
        ('site', cfg.get('site') or ''),
        ('unlimited_capacity_machine', ','.join(cfg.get('unlimited_capacity_machine') or [])),
        ('forecast_actuals_months', cfg.get('forecast_actuals_months')),
        ('forecast_align_to_month', bool(cfg.get('forecast_align_to_month', True))),
        ('purchased_and_produced', ','.join(
            f'{m}:{v}' for m, v in (cfg.get('purchased_and_produced') or {}).items())),
    ])

    fte = master.get('fte') or {}
    fte_rows = [
        ('fte_hours_per_year', fte.get('fte_hours_per_year')),
        ('default_shift_name', fte.get('default_shift_name') or ''),
    ]
    for shift_name, hours in sorted((fte.get('shift_hours') or {}).items()):
        fte_rows.append((f'shift_hours.{shift_name}', hours))
    params = fte.get('params') or {}
    for name in sorted(set(FTE_PARAM_DEFAULTS) | set(params)):
        fte_rows.append((f'{_FTE_PARAM_PREFIX}{name}',
                         params.get(name, FTE_PARAM_DEFAULTS.get(name))))
    _write_kv(wb, SHEET_FTE, fte_rows)

    _write_table(wb, SHEET_MATERIALS, _field_names(Material),
                 master.get('materials') or [])

    # shift_system is DERIVED state (finalize_shift_systems zet hem bij elke
    # load op UNLIMITED/3-shift o.b.v. de unlimited-lijst); hem exporteren
    # suggereerde een bewerkbaar veld waarvan edits stil verdwenen.
    machine_fields = [f for f in _field_names(Machine)
                      if f not in ('availability_by_period', 'shift_system')]
    periods = sorted({p for m in (master.get('machines') or [])
                      for p in (m.get('availability_by_period') or {})})
    machine_headers = machine_fields + [f'{_AVAILABILITY_PREFIX}{p}' for p in periods]
    machine_rows = []
    for m in master.get('machines') or []:
        row = {f: m.get(f) for f in machine_fields}
        for p in periods:
            row[f'{_AVAILABILITY_PREFIX}{p}'] = (m.get('availability_by_period') or {}).get(p)
        machine_rows.append(row)
    _write_table(wb, SHEET_MACHINES, machine_headers, machine_rows)

    _write_keyed_table(wb, SHEET_SAFETY, _field_names(SafetyStockConfig),
                       master.get('safety_stock') or {})

    purchase = master.get('purchase') or {}
    materials = sorted(set(purchase.get('lead_times') or {})
                       | set(purchase.get('moq') or {})
                       | set(purchase.get('sheet_materials') or []))
    _write_table(wb, SHEET_PURCHASE,
                 ['material', 'lead_time', 'moq', 'in_purchase_sheet'],
                 [{
                     'material': m,
                     'lead_time': (purchase.get('lead_times') or {}).get(m),
                     'moq': (purchase.get('moq') or {}).get(m),
                     'in_purchase_sheet': m in set(purchase.get('sheet_materials') or []),
                 } for m in materials])

    _write_keyed_table(wb, SHEET_PRICES, _field_names(SalesPriceItem),
                       master.get('sales_prices') or {})
    _write_keyed_table(wb, SHEET_MATERIAL_COSTS, _field_names(RawMaterialCost),
                       master.get('material_costs') or {})
    _write_keyed_table(wb, SHEET_MACHINE_COSTS, _field_names(MachineCost),
                       master.get('machine_costs') or {})

    vp = master.get('valuation_params') or {}
    _write_kv(wb, SHEET_VALUATION,
              [(name, vp.get(name)) for name in _field_names(ValuationParameters)])

    _write_fte_dataset_sheets(wb, master)

    path = Path(path)
    tmp = path.with_name(path.name + '.tmp')
    wb.save(str(tmp))
    tmp.replace(path)


def _write_kv(wb, sheet, rows):
    ws = wb.create_sheet(sheet)
    ws.append(['naam', 'waarde'])
    for key, value in rows:
        ws.append([key, value])


def _write_table(wb, sheet, headers, rows):
    ws = wb.create_sheet(sheet)
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(h) for h in headers])


def _write_keyed_table(wb, sheet, fields, keyed: dict):
    headers = [_KEY_COLUMN] + list(fields)
    ws = wb.create_sheet(sheet)
    ws.append(headers)
    for key in sorted(keyed):
        item = keyed[key]
        ws.append([key] + [_cell_value(item.get(f)) for f in fields])


def _cell_value(value):
    """Excel-safe rendering of the container fields the F2-CF datasets carry.

    A machine list becomes 'PBA01,PBA02'; a per-machine factor map becomes
    'PBA01:0.8,PBA02:1'. Everything else passes through untouched.
    """
    if isinstance(value, (list, tuple, set)):
        return ','.join(str(v) for v in value)
    if isinstance(value, dict):
        return ','.join(f'{k}:{v}' for k, v in value.items())
    return value


def _write_fte_dataset_sheets(wb, master: dict) -> None:
    for name, dc in FTE_DATASETS.items():
        _write_keyed_table(wb, FTE_DATASET_SHEETS[name], _field_names(dc),
                           master.get(name) or {})


# ------------------------------------------------------------------- parse


class MasterWorkbookError(ValueError):
    """Structured parse/validation failure with a user-facing message."""


def parse_master_workbook(path) -> Tuple[dict, dict]:
    """Parse an (edited) master workbook back into (master, meta).

    Purchase actuals are intentionally absent from the workbook; the caller
    merges the current store's actuals back in. Raises MasterWorkbookError
    with a Dutch, user-facing message on structural problems.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise MasterWorkbookError(f'Werkboek niet leesbaar: {exc}') from exc
    try:
        sheets = {ws.title: ws for ws in wb.worksheets}
        missing = [s for s in (SHEET_META, SHEET_CONFIG, SHEET_FTE, SHEET_MATERIALS,
                               SHEET_MACHINES, SHEET_SAFETY, SHEET_PURCHASE,
                               SHEET_PRICES, SHEET_MATERIAL_COSTS,
                               SHEET_MACHINE_COSTS, SHEET_VALUATION)
                   if s not in sheets]
        if missing:
            raise MasterWorkbookError(
                'Dit is geen masterwerkboek van de app: ontbrekende bladen '
                + ', '.join(missing))

        meta = dict(_read_kv(sheets[SHEET_META]))
        cfg_raw = dict(_read_kv(sheets[SHEET_CONFIG]))
        fte_raw = dict(_read_kv(sheets[SHEET_FTE]))
        vp_raw = dict(_read_kv(sheets[SHEET_VALUATION]))

        def guarded(sheet, fn):
            # Een tikfout als '10 ton' in een getalcel moet een nette
            # Nederlandse afwijzing geven, geen HTTP 500 met traceback.
            try:
                return fn()
            except MasterWorkbookError:
                raise
            except (ValueError, TypeError) as exc:
                raise MasterWorkbookError(
                    f'Ongeldige celwaarde in blad "{sheet}": {exc}') from exc

        master = {
            'schema_version': 1,
            'config': guarded(SHEET_CONFIG, lambda: _parse_config(cfg_raw)),
            'fte': guarded(SHEET_FTE, lambda: _parse_fte(fte_raw)),
            'materials': guarded(SHEET_MATERIALS, lambda: _parse_table(
                sheets[SHEET_MATERIALS], Material, require='material_number')),
            'machines': guarded(SHEET_MACHINES, lambda: _parse_machines(sheets[SHEET_MACHINES])),
            'safety_stock': guarded(SHEET_SAFETY, lambda: _parse_keyed_table(
                sheets[SHEET_SAFETY], SafetyStockConfig)),
            'purchase': guarded(SHEET_PURCHASE, lambda: _parse_purchase(sheets[SHEET_PURCHASE])),
            'sales_prices': guarded(SHEET_PRICES, lambda: _parse_keyed_table(
                sheets[SHEET_PRICES], SalesPriceItem)),
            'material_costs': guarded(SHEET_MATERIAL_COSTS, lambda: _parse_keyed_table(
                sheets[SHEET_MATERIAL_COSTS], RawMaterialCost)),
            'machine_costs': guarded(SHEET_MACHINE_COSTS, lambda: _parse_keyed_table(
                sheets[SHEET_MACHINE_COSTS], MachineCost)),
            'valuation_params': guarded(SHEET_VALUATION, lambda: _parse_valuation(vp_raw)),
        }
        # F2-CF sheets are optional: a workbook exported before they existed
        # must still import, and an absent sheet means "leave this dataset as
        # the store has it" — the import route merges, it does not clear.
        for name, dc in FTE_DATASETS.items():
            sheet = FTE_DATASET_SHEETS[name]
            if sheet in sheets:
                master[name] = guarded(
                    sheet, lambda ws=sheets[sheet], dc=dc: _parse_keyed_table(ws, dc))
        return master, meta
    finally:
        wb.close()


def _read_kv(ws):
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header
    for row in rows:
        if not row or row[0] is None:
            continue
        yield str(row[0]).strip(), (row[1] if len(row) > 1 else None)


def _read_table(ws, with_row_numbers: bool = False):
    """Rijen van een tabelblad als dict. Met with_row_numbers komt het ECHTE
    bladrijnummer mee: lege scheidingsregels worden overgeslagen, dus de
    positie in de gefilterde stroom loopt niet meer gelijk met wat de
    gebruiker in Excel ziet — en een foutmelding die de verkeerde rij aanwijst
    is erger dan geen rijnummer."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None) or ()
    headers = [str(h).strip() if h is not None else '' for h in header]
    for number, row in enumerate(rows, start=2):
        if _is_empty_row(row):
            continue
        yield (number, _row_dict(headers, row)) if with_row_numbers else _row_dict(headers, row)


def _parse_pap_entries(raw_value) -> dict:
    """Parse 'MAT:fractie, MAT:fractie' met NL-tolerantie en harde afwijzing.

    Een decimale komma in de fractie ('MAT1:0,45') splitst op het eerste
    gezicht in twee fragmenten; fragmenten zonder ':' worden daarom weer aan
    hun voorganger geplakt. Malformed entries of fracties buiten [0, 1] geven
    een MasterWorkbookError — stil overslaan maakte van een tikfout een
    fractie 0,0 (= stil 100% inkoop) of liet de materiaalsplit verdwijnen."""
    text = str(raw_value or '').strip()
    if not text:
        return {}
    fragments = [f.strip() for f in text.split(',')]
    entries = []
    for frag in fragments:
        if not frag:
            continue
        if ':' not in frag and entries:
            entries[-1] += ',' + frag  # decimale komma weer aan elkaar
        else:
            entries.append(frag)
    pap = {}
    for entry in entries:
        parts = entry.split(':')
        material = parts[0].strip() if parts else ''
        if len(parts) != 2 or not material:
            raise MasterWorkbookError(
                f'Config: purchased_and_produced-onderdeel "{entry}" is geen '
                f'"MATERIAAL:fractie"-paar.')
        try:
            fraction = float(parts[1].strip().replace(',', '.'))
        except ValueError:
            raise MasterWorkbookError(
                f'Config: productiefractie bij {material} is geen getal '
                f'("{parts[1].strip()}").')
        if not 0.0 <= fraction <= 1.0:
            raise MasterWorkbookError(
                f'Config: productiefractie bij {material} moet tussen 0 en 1 '
                f'liggen (was {fraction:g}).')
        pap[material] = fraction
    return pap


def _parse_config(raw: dict) -> dict:
    pap = _parse_pap_entries(raw.get('purchased_and_produced'))
    initial_date = raw.get('initial_date')
    if isinstance(initial_date, datetime):
        initial_date = initial_date.isoformat()
    if _empty(initial_date):
        # Stil terugvallen op het hydratie-anker (dec 2025) zou alle
        # periodesleutels verschuiven zonder dat iemand het merkt.
        raise MasterWorkbookError('Config: initial_date mag niet leeg zijn.')
    align_raw = raw.get('forecast_align_to_month')
    return {
        'initial_date': str(initial_date),
        'forecast_months': int(float(raw.get('forecast_months') or 12)),
        'site': str(raw.get('site') or '').strip(),
        'unlimited_capacity_machine': [
            m.strip() for m in str(raw.get('unlimited_capacity_machine') or '').split(',')
            if m.strip()],
        'forecast_actuals_months': int(float(raw.get('forecast_actuals_months') or 12)),
        # Lege cel betekent "default" (True), net als elke andere lege cel —
        # _to_bool(None) zou stil naar de positionele validatiemodus wisselen.
        'forecast_align_to_month': True if _empty(align_raw) else _to_bool(align_raw),
        'purchased_and_produced': pap,
    }


def _parse_fte(raw: dict) -> dict:
    shift_hours = {}
    params = dict(FTE_PARAM_DEFAULTS)
    for key, value in raw.items():
        if key.startswith('shift_hours.') and value is not None:
            shift_hours[key[len('shift_hours.'):]] = float(value)
        elif key.startswith(_FTE_PARAM_PREFIX) and value is not None:
            params[key[len(_FTE_PARAM_PREFIX):]] = float(value)
    return {
        'fte_hours_per_year': float(raw.get('fte_hours_per_year') or 1492),
        'shift_hours': shift_hours,
        'default_shift_name': str(raw.get('default_shift_name') or '3-shift system'),
        'params': params,
    }


def _empty(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _parse_table(ws, dc, require: Optional[str] = None) -> list:
    casters = _casters(dc)
    fields = _field_names(dc)
    items = []
    for row in _read_table(ws):
        # Een half-gewiste rij (identiteit leeg, één cel achtergebleven) mag
        # nooit als record met lege sleutel de store in.
        if require and _empty(row.get(require)):
            continue
        item = {}
        for f in fields:
            item[f] = casters[f](row.get(f)) if f in casters else row.get(f)
        items.append(item)
    return items


def _parse_machines(ws) -> list:
    casters = _casters(Machine)
    fields = [f for f in _field_names(Machine)
              if f not in ('availability_by_period', 'shift_system')]
    items = []
    for row in _read_table(ws):
        if _empty(row.get('machine_code')):
            continue
        # OEE leeg → 0.0 zou stil alle capaciteit van de machine wissen.
        if _empty(row.get('oee')):
            raise MasterWorkbookError(
                f"Machine {row.get('machine_code')}: OEE mag niet leeg zijn.")
        item = {f: casters[f](row.get(f)) for f in fields}
        # shift_system is afgeleid (finalize_shift_systems) en staat niet in
        # het werkboek; placeholder — de importroute neemt de opgeslagen
        # waarde over, hydratie herleidt hem daarna hoe dan ook.
        item['shift_system'] = '3-shift system'
        availability = {}
        for col, value in row.items():
            if col.startswith(_AVAILABILITY_PREFIX) and value is not None:
                availability[col[len(_AVAILABILITY_PREFIX):].strip()] = float(value)
        item['availability_by_period'] = availability
        items.append(item)
    return items


def _parse_keyed_table(ws, dc) -> dict:
    """Blad met een sleutelkolom → {sleutel: record}.

    Twee rijen met dezelfde sleutel zijn een fout, geen keuze: de tweede
    overschreef de eerste stil, dus de helft van een gekopieerde regel
    verdween zonder dat de importdiff iets liet zien. Bij handmatig bewerkte
    werkboeken is dupliceren juist de normale manier om een regel toe te
    voegen — en dan de sleutel vergeten aan te passen de normale vergissing.
    """
    casters = _casters(dc)
    fields = _field_names(dc)
    keyed = {}
    seen_rows = {}
    for number, row in _read_table(ws, with_row_numbers=True):
        key = row.get(_KEY_COLUMN)
        if key is None or str(key).strip() == '':
            continue
        key = str(key).strip()
        if key in keyed:
            raise MasterWorkbookError(
                f'Blad "{ws.title}": sleutel "{key}" staat twee keer '
                f'(rij {seen_rows[key]} en rij {number}). Geef elke rij een '
                f'eigen sleutel — anders verdwijnt er stil één.')
        seen_rows[key] = number
        keyed[key] = {f: casters[f](row.get(f)) for f in fields}
    return keyed


def _parse_purchase(ws) -> dict:
    lead_times, moq, sheet_materials = {}, {}, []
    for row in _read_table(ws):
        material = str(row.get('material') or '').strip()
        if not material:
            continue
        if row.get('lead_time') is not None:
            lead_times[material] = int(float(row['lead_time']))
        if row.get('moq') is not None:
            moq[material] = float(row['moq'])
        if _to_bool(row.get('in_purchase_sheet')):
            sheet_materials.append(material)
    return {
        'lead_times': lead_times,
        'moq': moq,
        'sheet_materials': sorted(sheet_materials),
        # Actuals zijn maanddata en staan bewust niet in het werkboek; de
        # importroute behoudt de actuals van de huidige store.
        'actuals': {},
    }


# ------------------------------------------------------------ equivalents


_IDENTITY_KEYS = ('material_number', 'machine_code')

_ISO_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}')


def _identity(item):
    if isinstance(item, dict):
        for key in _IDENTITY_KEYS:
            if key in item:
                return item.get(key)
    return None


def absorb_equivalents(previous, incoming):
    """Replace semantically-equal values in `incoming` by the exact value
    from `previous` (mutates and returns `incoming`).

    Excel stores floats at 15 significant digits and renders '' and None
    identically, so a pure export->parse round trip drifts in the 16th
    digit and in empty-string-vs-None. Without this step every import diff
    would flag those as changes and a confirmed no-op import would churn
    the store. Genuine edits pass through untouched."""
    if isinstance(incoming, dict) and isinstance(previous, dict):
        for key, value in incoming.items():
            if key in previous:
                incoming[key] = absorb_equivalents(previous[key], value)
        return incoming
    if isinstance(incoming, list) and isinstance(previous, list):
        prev_by_id = {ident: p for p in previous
                      if (ident := _identity(p)) is not None}
        for i, item in enumerate(incoming):
            ident = _identity(item)
            if ident is not None and ident in prev_by_id:
                incoming[i] = absorb_equivalents(prev_by_id[ident], item)
        return incoming
    if incoming == previous:
        return incoming
    if incoming in (None, '') and previous in (None, ''):
        return previous
    if (isinstance(incoming, (int, float)) and isinstance(previous, (int, float))
            and not isinstance(incoming, bool) and not isinstance(previous, bool)
            and math.isclose(float(incoming), float(previous),
                             rel_tol=1e-9, abs_tol=1e-12)):
        return previous
    if (isinstance(incoming, str) and isinstance(previous, str)
            and _ISO_DATE_RE.match(incoming) and _ISO_DATE_RE.match(previous)
            and incoming[:10] == previous[:10]
            and incoming[10:] in ('', 'T00:00:00')
            and previous[10:] in ('', 'T00:00:00')):
        # '2026-01-01' vs '2026-01-01T00:00:00': Excel maakt van een opnieuw
        # bevestigde datumcel een datetime; semantisch hetzelfde moment, dus
        # geen diff en geen representatie-churn in de store.
        return previous
    return incoming


def _parse_valuation(raw: dict) -> Optional[dict]:
    values = {}
    for name in _field_names(ValuationParameters):
        value = raw.get(name)
        if value is None:
            continue
        values[name] = int(float(value)) if name.startswith('days_') else float(value)
    return values or None
