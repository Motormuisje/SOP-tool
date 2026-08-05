"""F2-CF masterdata: bemensing, loonkosten, combinaties, indirecte
activiteiten en doorzet-overrides door serialize/hydrate/overlay, het
masterwerkboek en de PATCH-route."""

import json
from types import SimpleNamespace

import pytest
from flask import Flask

from modules.master_data import (
    FTE_DATASETS,
    hydrate_loader,
    overlay_master_data,
    serialize_master,
)
from modules.master_workbook import (
    FTE_DATASET_SHEETS,
    export_master_workbook,
    parse_master_workbook,
)
from modules.models import FTE_PARAM_DEFAULTS, derive_effective_fte_hours
from tests.master_fixtures import fake_master_loader
from ui import master_store
from ui.routes.master_data import create_master_data_blueprint

pytestmark = pytest.mark.no_fixture


def _master():
    return json.loads(json.dumps(serialize_master(fake_master_loader()), default=str))


class _Loader:
    """Kaal hydratatiedoel — hydrate_loader zet alle attributen zelf."""


def test_serialize_carries_every_fte_dataset():
    master = _master()
    for name in FTE_DATASETS:
        assert name in master, f'{name} ontbreekt in de geserialiseerde master'
    assert master['fte']['params']['utilization_rate'] == 0.85
    assert master['staffing_norms']['G1']['operators_per_hour'] == 1.5


def test_hydrate_round_trip_gives_dataclasses():
    loader = _Loader()
    hydrate_loader(loader, _master())

    assert loader.staffing_norms['G1'].operators_per_hour == 1.5
    assert loader.staffing_norms['G1'].scope == 'group'
    assert loader.labor_rates['operators'].cost_per_fte_per_year == 58000.0
    combo = loader.machine_combinations['C1']
    assert combo.machine_codes == ['PBA01', 'PBA02']
    assert combo.factor_for('PBA02') == 0.75      # per-machine override
    assert combo.factor_for('PBA01') == 0.9       # combinatiebrede factor
    assert loader.indirect_activities['TRUCK'].tons_per_truck == 25.0
    assert loader.throughput_overrides['PBA01|M1'].source == 'MES'
    assert loader.benchmark_throughput['PBA01|M1'].mes_t_per_hour == 27.0
    assert loader.fte_params['illness_pct'] == 0.10


def test_hydrate_defaults_when_datasets_absent():
    """Een store van vóór F2-CF mag niet crashen: lege datasets, defaults."""
    master = _master()
    for name in FTE_DATASETS:
        master.pop(name)
    master['fte'].pop('params')

    loader = _Loader()
    hydrate_loader(loader, master)

    assert loader.staffing_norms == {}
    assert loader.machine_combinations == {}
    assert loader.fte_params == FTE_PARAM_DEFAULTS


def test_hydrate_rejects_non_numeric_norm():
    master = _master()
    master['staffing_norms']['G1']['operators_per_hour'] = 'twee'
    with pytest.raises(ValueError, match='StaffingNorm'):
        hydrate_loader(_Loader(), master)


def test_overlay_merges_onto_workbook_loader():
    """Werkboeksessie: de store vult de F2-CF-datasets aan, wist niets."""
    loader = _Loader()
    hydrate_loader(loader, _master())
    loader._apply_pap_override = lambda replace=False: None
    loader._extend_machine_availability_to_periods = lambda: None
    existing = loader.staffing_norms['G1']

    incoming = _master()
    incoming['staffing_norms'] = {
        'G2': {'code': 'G2', 'operators_per_hour': 2.0, 'scope': 'group',
               'function_group': '', 'description': ''}}
    overlay_master_data(loader, incoming)

    assert loader.staffing_norms['G2'].operators_per_hour == 2.0
    assert loader.staffing_norms['G1'] is existing  # bestaande norm ongemoeid


def test_workbook_round_trip_keeps_lists_and_maps(tmp_path):
    master = _master()
    path = tmp_path / 'master.xlsx'
    export_master_workbook(master, path, site='NLX1', store_version=3)
    parsed, _ = parse_master_workbook(path)

    combo = parsed['machine_combinations']['C1']
    assert combo['machine_codes'] == ['PBA01', 'PBA02']
    assert combo['throughput_factor_by_machine'] == {'PBA02': 0.75}
    assert combo['is_active'] is True
    assert parsed['fte']['params']['illness_pct'] == 0.10
    for name in FTE_DATASETS:
        assert parsed[name] == master[name], f'werkboek-round-trip wijkt af in {name}'


def test_workbook_without_fte_sheets_still_parses(tmp_path):
    """Werkboek van vóór F2-CF: de bladen ontbreken, de import moet lukken."""
    import openpyxl

    master = _master()
    path = tmp_path / 'oud.xlsx'
    export_master_workbook(master, path, site='NLX1', store_version=3)
    wb = openpyxl.load_workbook(str(path))
    for sheet in FTE_DATASET_SHEETS.values():
        wb.remove(wb[sheet])
    wb.save(str(path))

    parsed, _ = parse_master_workbook(path)
    for name in FTE_DATASETS:
        assert name not in parsed  # afwezig ≠ leeg: de route behoudt de store


def test_workbook_rejects_a_duplicate_key(tmp_path):
    """Een regel dupliceren is de normale manier om er in Excel één bij te
    maken; de sleutel vergeten aan te passen de normale vergissing. De tweede
    rij overschreef de eerste stil, dus verdween er een norm zonder dat de
    importdiff iets liet zien."""
    import openpyxl

    from modules.master_workbook import MasterWorkbookError

    master = _master()
    path = tmp_path / 'master.xlsx'
    export_master_workbook(master, path, site='NLX1', store_version=3)
    wb = openpyxl.load_workbook(str(path))
    ws = wb[FTE_DATASET_SHEETS['staffing_norms']]
    ws.append([ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)])
    wb.save(str(path))

    with pytest.raises(MasterWorkbookError, match='twee keer'):
        parse_master_workbook(path)


def test_workbook_rejects_malformed_factor_map(tmp_path):
    import openpyxl

    from modules.master_workbook import MasterWorkbookError

    master = _master()
    path = tmp_path / 'master.xlsx'
    export_master_workbook(master, path, site='NLX1', store_version=3)
    wb = openpyxl.load_workbook(str(path))
    ws = wb[FTE_DATASET_SHEETS['machine_combinations']]
    headers = [c.value for c in ws[1]]
    ws.cell(row=2, column=headers.index('throughput_factor_by_machine') + 1,
            value='PBA02')  # geen SLEUTEL:waarde
    wb.save(str(path))

    with pytest.raises(MasterWorkbookError):
        parse_master_workbook(path)


def test_derive_effective_fte_hours_matches_client_model():
    """Maastricht: 2080 − 224 − 112 − 48 = 1696; −169,60 −33,92 = 1492,48.

    Beide percentages gaan over DEZELFDE basis. Zou de afleiding ze stapelen
    (×0,90 ×0,98) dan kwam er 1495,87 uit — 3,4 uur per FTE per jaar te veel,
    en dus systematisch te weinig benodigde FTE.
    """
    params = fake_master_loader().fte_params
    hours = derive_effective_fte_hours(params)
    assert hours == pytest.approx(1492.48, abs=0.01)

    base = 2080 - 224 - 112 - 48
    assert hours == pytest.approx(base - base * 0.10 - base * 0.02)
    assert hours != pytest.approx(base * 0.90 * 0.98)


# ── routes ──────────────────────────────────────────────────────────────────


@pytest.fixture
def md_app(tmp_path):
    previous = master_store.get_store_path()
    path = tmp_path / 'master_store.json'
    master_store.set_store_path(path)
    master_store.save_master_store(path, _master(), source_filename='seed.xlsm')
    master_store.set_store_path(path)
    app = Flask(__name__)
    app.config['TESTING'] = True
    app.register_blueprint(create_master_data_blueprint({}, lambda: tmp_path / 'uploads'))
    yield SimpleNamespace(client=app.test_client(), path=path)
    master_store.set_store_path(previous or path)


def test_patch_staffing_norms(md_app):
    body = md_app.client.get('/api/master_data/staffing_norms').get_json()
    value = body['value']
    value['G2'] = {'code': 'G2', 'operators_per_hour': 2.5, 'scope': 'group',
                   'function_group': 'operators', 'description': ''}
    res = md_app.client.patch('/api/master_data/staffing_norms', json={'value': value})
    assert res.status_code == 200
    stored = master_store.get_current_master_record()['master']['staffing_norms']
    assert stored['G2']['operators_per_hour'] == 2.5


def test_patch_rejects_invalid_norm(md_app):
    res = md_app.client.patch('/api/master_data/staffing_norms', json={'value': {
        'G2': {'code': 'G2', 'operators_per_hour': 'veel'}}})
    assert res.status_code == 400
    assert 'geweigerd' in res.get_json()['error']


class TestKeyMatchesIdentity:
    """De motor zoekt records op met hun DICT-SLEUTEL; de identiteitsvelden
    staan er als bewerkbare kolommen naast. Liepen die uiteen, dan werkte een
    doorzet-override nog steeds op de machine uit de sleutel terwijl het record
    een andere machine noemde — zonder enige melding."""

    def test_conflicting_identity_field_is_refused(self):
        master = _master()
        master['throughput_overrides']['PBA01|M1']['machine_code'] = 'PBA02'
        with pytest.raises(ValueError, match='PBA01\\|M1'):
            hydrate_loader(_Loader(), master)

    def test_empty_identity_field_is_filled_from_the_key(self):
        """De veelvoorkomende typfout: alleen de sleutelkolom ingevuld."""
        master = _master()
        master['throughput_overrides']['PBA01|M1']['machine_code'] = ''
        master['throughput_overrides']['PBA01|M1']['material_number'] = ''
        loader = _Loader()
        hydrate_loader(loader, master)

        override = loader.throughput_overrides['PBA01|M1']
        assert override.machine_code == 'PBA01'
        assert override.material_number == 'M1'

    def test_machine_level_benchmark_key_keeps_its_empty_half(self):
        master = _master()
        master['benchmark_throughput']['PBA07|'] = {
            'machine_code': '', 'material_number': '', 'mes_t_per_hour': 0.0,
            'peer_t_per_hour': 6.21, 'mes_oee': 0.0, 'note': ''}
        loader = _Loader()
        hydrate_loader(loader, master)

        entry = loader.benchmark_throughput['PBA07|']
        assert entry.machine_code == 'PBA07'
        assert entry.material_number == ''

    def test_wrong_key_shape_is_refused(self):
        master = _master()
        master['throughput_overrides']['PBA01'] = dict(
            master['throughput_overrides']['PBA01|M1'])
        with pytest.raises(ValueError, match='vorm'):
            hydrate_loader(_Loader(), master)

    @pytest.mark.parametrize('dataset,key,record', [
        ('labor_rates', 'Operatie | Maalderij',
         {'function_group': 'Operatie | Maalderij', 'cost_per_fte_per_year': 1.0}),
        ('indirect_activities', 'Laden|Lossen', {'activity_id': 'Laden|Lossen'}),
        ('machine_combinations', 'MILL|SIEVE', {'combination_id': 'MILL|SIEVE'}),
        ('staffing_norms', 'ZZ|G1', {'code': 'ZZ|G1', 'operators_per_hour': 1.0}),
    ])
    def test_a_pipe_in_a_free_text_identity_is_allowed(self, dataset, key, record):
        """Deze datasets hebben ÉÉN identiteitsveld en dat is vrije tekst — de
        toevoegknop zegt letterlijk 'Functiegroep'. Onvoorwaardelijk op '|'
        splitsen maakte zo'n sleutel ongeldig en daarmee de HELE store
        onlaadbaar: niet alleen de werkbank, maar elke berekening."""
        master = _master()
        master[dataset] = {key: record}

        loader = _Loader()
        hydrate_loader(loader, master)

        assert key in getattr(loader, dataset)

    def test_a_pair_key_still_needs_both_halves(self):
        """Bij twee identiteitsvelden blijft de vorm wél afgedwongen."""
        master = _master()
        master['throughput_overrides'] = {'ALLEEN_MACHINE': {
            'machine_code': 'M', 'material_number': 'P',
            'throughput_t_per_hour': 1.0}}
        with pytest.raises(ValueError, match='vorm'):
            hydrate_loader(_Loader(), master)

    def test_a_material_number_containing_a_pipe_lands_in_the_last_half(self):
        """Met maxsplit blijft een '|' in het materiaalnummer bij het
        materiaal horen in plaats van de sleutelvorm te breken."""
        master = _master()
        master['throughput_overrides'] = {'PBA01|M|1': {
            'machine_code': '', 'material_number': '',
            'throughput_t_per_hour': 1.0}}

        loader = _Loader()
        hydrate_loader(loader, master)

        entry = loader.throughput_overrides['PBA01|M|1']
        assert entry.machine_code == 'PBA01'
        assert entry.material_number == 'M|1'

    def test_a_key_with_surrounding_spaces_is_normalised(self):
        """De sleutel werd gestript vergeleken maar ONgestript opgeslagen, dus
        de motor zocht op 'ZZ_GROUP01' en vond ' ZZ_GROUP01 ' niet: de norm
        werd stil genegeerd en de werkbank viel terug op de L12-coëfficiënt."""
        master = _master()
        master['staffing_norms'] = {' ZZ_GROUP01 ': {
            'code': 'ZZ_GROUP01', 'operators_per_hour': 2.0, 'scope': 'group'}}

        loader = _Loader()
        hydrate_loader(loader, master)

        assert 'ZZ_GROUP01' in loader.staffing_norms
        assert loader.staffing_norms['ZZ_GROUP01'].operators_per_hour == 2.0

    def test_combination_key_must_match_its_id(self):
        master = _master()
        master['machine_combinations']['C2'] = dict(master['machine_combinations']['C1'])
        with pytest.raises(ValueError, match='C2'):
            hydrate_loader(_Loader(), master)


class TestDatasetValueValidation:
    """Waarden die alleen door hydratie gingen: type klopte, betekenis niet."""

    @pytest.mark.parametrize('dataset,value,fragment', [
        ('staffing_norms',
         {'G1': {'code': 'G1', 'operators_per_hour': 1.0, 'scope': 'groep'}},
         'bereik'),
        ('staffing_norms',
         {'G1': {'code': 'G1', 'operators_per_hour': -1.0}},
         '0 of hoger'),
        ('indirect_activities',
         {'X': {'activity_id': 'X', 'driver': 'per_moon'}},
         'driver'),
        ('throughput_overrides',
         {'M|P': {'machine_code': 'M', 'material_number': 'P',
                  'throughput_t_per_hour': 0.0}},
         'groter dan 0'),
        ('machine_combinations',
         {'C': {'combination_id': 'C', 'throughput_factor': 0.0}},
         'doorzetfactor'),
    ])
    def test_nonsense_is_refused_with_a_readable_reason(self, md_app, dataset, value, fragment):
        res = md_app.client.patch(f'/api/master_data/{dataset}', json={'value': value})
        assert res.status_code == 400
        assert fragment in res.get_json()['error']

    def test_a_driver_without_its_divider_warns_instead_of_failing(self, md_app):
        value = {'T': {'activity_id': 'T', 'driver': 'per_truck',
                       'hours_per_unit': 0.75, 'tons_per_truck': 0.0}}
        res = md_app.client.patch('/api/master_data/indirect_activities',
                                  json={'value': value})
        assert res.status_code == 200
        assert any('ton per truck' in w for w in res.get_json().get('warnings', []))


def test_status_counts_include_fte_datasets(md_app):
    counts = md_app.client.get('/api/master_data').get_json()['counts']
    assert counts['staffing_norms'] == 1
    assert counts['indirect_activities'] == 3
