"""Masterwerkboek (per-site Excel-bewerkingsmedium): round-trip-garantie,
bewerkingen, en structurele validatie."""

import json

import pytest

from modules.master_data import serialize_master
from modules.master_workbook import (
    MasterWorkbookError,
    export_master_workbook,
    parse_master_workbook,
)
from tests.master_fixtures import fake_master_loader

pytestmark = pytest.mark.no_fixture


def _master():
    return json.loads(json.dumps(serialize_master(fake_master_loader()), default=str))


def _export(tmp_path, master, site='NLX1', version=5):
    path = tmp_path / 'SOP_Masterdata_NLX1.xlsx'
    export_master_workbook(master, path, site=site, store_version=version)
    return path


def test_export_parse_round_trip_is_identical(tmp_path):
    master = _master()
    path = _export(tmp_path, master)
    parsed, meta = parse_master_workbook(path)

    assert meta['site'] == 'NLX1'
    assert int(meta['store_version']) == 5
    # Actuals staan bewust niet in het werkboek (maanddata); de rest moet
    # byte-voor-byte terugkomen.
    expected = json.loads(json.dumps(master))
    expected['purchase']['actuals'] = {}
    parsed_norm = json.loads(json.dumps(parsed))
    for key in expected:
        if key == 'schema_version':
            continue
        assert parsed_norm[key] == expected[key], f'round-trip afwijking in {key}'


def test_edits_in_excel_come_back(tmp_path):
    master = _master()
    path = _export(tmp_path, master)

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb['Materialen']
    headers = [c.value for c in ws[1]]
    name_col = headers.index('name') + 1
    active_col = headers.index('is_active') + 1
    ws.cell(row=2, column=name_col, value='HERNOEMD IN EXCEL')
    ws.cell(row=3, column=active_col, value=False)
    # nieuw materiaal onderaan
    new_row = [None] * len(headers)
    new_row[headers.index('material_number')] = 'M9'
    new_row[headers.index('name')] = 'Nieuw product'
    new_row[headers.index('product_type')] = 'Raw Material'
    new_row[headers.index('product_family')] = 'FAM'
    ws.append(new_row)
    wb.save(str(path))

    parsed, _ = parse_master_workbook(path)
    by_num = {m['material_number']: m for m in parsed['materials']}
    assert by_num['M1']['name'] == 'HERNOEMD IN EXCEL'
    assert by_num['M2']['is_active'] is False
    assert by_num['M9']['product_type'] == 'Raw Material'
    assert by_num['M9']['is_active'] is True  # lege cel → default actief


def test_machine_availability_grid_round_trips(tmp_path):
    master = _master()
    master['machines'][0]['availability_by_period'] = {'2026-01': 0.5, '2026-02': 0.75}
    path = _export(tmp_path, master)
    parsed, _ = parse_master_workbook(path)
    assert parsed['machines'][0]['availability_by_period'] == {'2026-01': 0.5, '2026-02': 0.75}


def test_parse_refuses_foreign_workbook(tmp_path):
    import openpyxl
    path = tmp_path / 'vreemd.xlsx'
    wb = openpyxl.Workbook()
    wb.active.title = 'Blad1'
    wb.save(str(path))
    with pytest.raises(MasterWorkbookError, match='geen masterwerkboek'):
        parse_master_workbook(path)


def test_parsed_master_hydrates(tmp_path):
    """Validatie-door-hydratie: het geparste dict moet door dezelfde
    hydratie kunnen als de store zelf."""
    from modules.data_loader import DataLoader
    from modules.master_data import hydrate_loader

    master = _master()
    path = _export(tmp_path, master)
    parsed, _ = parse_master_workbook(path)
    parsed['purchase']['actuals'] = master['purchase']['actuals']

    import contextlib, io
    probe = DataLoader(master_data=parsed)
    with contextlib.redirect_stdout(io.StringIO()):
        hydrate_loader(probe, parsed)
    assert set(probe.materials) == {'M1', 'M2'}
    assert probe.machines['PBA01'].oee == 0.8


def test_absorb_equivalents_kills_excel_drift_but_keeps_real_edits():
    from modules.master_workbook import absorb_equivalents
    previous = {
        'materials': [
            {'material_number': 'M1', 'name': 'A', 'spc_product': '', 'fte_requirements': 1175.3867293132146},
            {'material_number': 'M2', 'name': 'B', 'spc_product': None, 'fte_requirements': 2.0},
        ],
        'safety_stock': {'M1': {'safety_stock': 0.010184287099903006}},
    }
    incoming = {
        'materials': [
            {'material_number': 'M1', 'name': 'A', 'spc_product': None, 'fte_requirements': 1175.386729313215},
            {'material_number': 'M2', 'name': 'ECHT GEWIJZIGD', 'spc_product': '', 'fte_requirements': 3.5},
        ],
        'safety_stock': {'M1': {'safety_stock': 0.01018428709990301}},
    }
    absorb_equivalents(previous, incoming)
    m1, m2 = incoming['materials']
    assert m1 == previous['materials'][0]                      # pure drift → exact terug
    assert m2['name'] == 'ECHT GEWIJZIGD' and m2['fte_requirements'] == 3.5  # echte edits blijven
    assert m2['spc_product'] == previous['materials'][1]['spc_product']      # leeg-equivalent
    assert incoming['safety_stock']['M1'] == previous['safety_stock']['M1']
