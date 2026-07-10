from datetime import datetime
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import modules.planning_engine as planning_module
from modules.models import (
    LineType,
    Machine,
    Material,
    PlanningConfig,
    PlanningRow,
    ProductType,
    RoutingItem,
)
from modules.planning_engine import PlanningEngine


pytestmark = pytest.mark.no_fixture


PERIODS = ["2025-01", "2025-02"]


def _row(material, line_type, values=None, aux=None, aux2=None):
    return PlanningRow(
        material_number=material,
        material_name=f"Name {material}",
        product_type="Bulk Product",
        product_family="Family",
        spc_product="SPC",
        product_cluster="Cluster",
        product_name="Product",
        line_type=line_type,
        aux_column=aux,
        aux_2_column=aux2,
        starting_stock=5.0,
        values=values or {"2025-01": 1.0, "2025-02": 2.0},
    )


def _engine_with_rows():
    engine = PlanningEngine(file_path="synthetic.xlsm", planning_month="2025-01")
    engine.data = SimpleNamespace(
        materials={"MAT-2": object(), "MAT-1": object()},
        bom=["bom-item"],
        machines={"M1": object()},
        machine_groups={"G1": object()},
        periods=PERIODS,
    )
    engine.results[LineType.DEMAND_FORECAST.value] = [
        _row("MAT-2", LineType.DEMAND_FORECAST.value, aux="2"),
        _row("MAT-1", LineType.DEMAND_FORECAST.value, aux="1"),
    ]
    engine.results[LineType.INVENTORY.value] = [
        _row("MAT-1", LineType.INVENTORY.value, values={"2025-01": -1.0, "2025-02": 3.0})
    ]
    engine.results[LineType.UTILIZATION_RATE.value] = [
        _row("Z_M1", LineType.UTILIZATION_RATE.value, values={"2025-01": 1.2}, aux="0.8")
    ]
    engine.value_results = {
        LineType.CONSOLIDATION.value: [
            _row("ZZZZZZ_TURNOVER", LineType.CONSOLIDATION.value, aux2="100")
        ]
    }
    return engine


def test_planning_engine_helpers_compile_summary_dataframe_and_json():
    engine = _engine_with_rows()

    engine._compile_all_rows()
    engine._generate_summary()
    df = engine.to_dataframe()
    payload = engine.to_json()

    assert engine.get_all_rows() == engine.all_rows
    assert engine.get_rows_by_type(LineType.INVENTORY.value)
    assert engine.get_summary()["total_rows"] == 4
    assert engine.get_summary()["period_list"] == PERIODS
    assert df["Material number"].tolist() == ["MAT-1", "MAT-1", "MAT-2", "Z_M1"]
    assert payload["summary"] == engine.summary
    assert payload["periods"] == PERIODS
    assert LineType.DEMAND_FORECAST.value in payload["results"]


def test_rebuild_machine_output_caches_handles_missing_data_and_bad_routing():
    engine = PlanningEngine()

    engine.rebuild_machine_output_caches()
    assert engine.machine_throughput_theo == {}
    assert engine.output_by_machine_period == {}

    machine = Machine("M1", "M1", "Machine 1", 1.0)
    routing = RoutingItem("P", "MAT-1", "Material", "M1", base_quantity=10.0, standard_time=2.0)
    bad_routing = RoutingItem("P", "MAT-2", "Material", "M2", base_quantity=0.0, standard_time=2.0)

    def get_all_routings(material):
        if material == "BROKEN":
            raise RuntimeError("routing lookup failed")
        if material == "MAT-1":
            return [routing]
        if material == "MAT-2":
            return [bad_routing]
        return []

    engine.data = SimpleNamespace(
        periods=PERIODS,
        materials={"MAT-1": object(), "MAT-2": object(), "BROKEN": object()},
        machines={"M1": machine},
        get_all_routings=get_all_routings,
    )
    engine.all_production_plans = {
        "MAT-1": {"2025-01": 20.0, "2025-02": 0.0},
        "BROKEN": {"2025-01": 999.0},
    }

    engine.rebuild_machine_output_caches()

    assert engine.machine_throughput_theo == {"M1": 5.0}
    assert engine.output_by_machine_period["M1"] == {"2025-01": 20.0, "2025-02": 0.0}


def test_to_excel_writes_planning_sheet(tmp_path):
    engine = _engine_with_rows()
    output = tmp_path / "planning.xlsx"

    engine.to_excel(str(output))

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Planning Results"]
    assert workbook["Planning Results"]["A1"].value == "Material number"


def test_excel_formatting_helpers_apply_expected_styles():
    engine = PlanningEngine()
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Material number",
            "Material name",
            "Line type",
            "Aux Column",
            "Aux 2 Column",
            "Starting stock",
            "2025-01",
        ]
    )
    ws.append(["MAT-1", "Name", LineType.TOTAL_DEMAND.value, 1, 2, 5, 10])
    ws.append(["MAT-1", "Name", LineType.INVENTORY.value, 1, 2, 5, -1])
    ws.append(["Z_M1", "Machine", LineType.UTILIZATION_RATE.value, 0.5, 1, 0, 1.2])
    ws.append(["M1", "Machine", LineType.AVAILABLE_CAPACITY.value, 1, 0.5, 0, 0.5])

    engine._apply_excel_formatting(ws)

    assert ws["A2"].font.bold is True
    assert ws["G4"].number_format == "0.0%"
    assert ws["E5"].number_format == "0%"
    assert ws["G1"].number_format == "mm/yyyy"
    assert len(ws.conditional_formatting) >= 3


def test_consolidation_and_fte_formatting_helpers():
    engine = PlanningEngine()
    wb = Workbook()
    ws = wb.active
    ws.append(["Material number", "Line type", "2025-01"])
    ws.append(["ZZZZZZ_TURNOVER", LineType.CONSOLIDATION.value, 100])
    ws.append(["ZZZZZZ_EBIT", LineType.CONSOLIDATION.value, 50])
    ws.append(["MAT-1", LineType.DEMAND_FORECAST.value, 1])

    engine._apply_consol_colors(ws)

    assert ws["A2"].fill.fill_type == "solid"
    assert ws["A3"].font.bold is True
    assert ws["A4"].fill.fill_type is None

    fte_ws = wb.create_sheet("FTE")
    fte_ws.append(["Material number", "Group name", "FTE needed", "2025-01", "Average"])
    fte_ws.append(["ZZ_GROUP", "Group", "", 1.23, 1.23])
    fte_ws.append(["TOTAL", "TOTAL", "", 2.0, 2.0])

    engine._apply_fte_formatting(fte_ws, ["2025-01"])

    assert fte_ws["A1"].font.bold is True
    assert fte_ws["D1"].number_format == "mm/yyyy"
    assert fte_ws["D2"].number_format == "#,##0.00"
    assert fte_ws["A3"].font.bold is True


def test_planning_engine_run_uses_orchestrators_in_order(monkeypatch):
    events = []

    class FakeDataLoader:
        def __init__(self, *args, **kwargs):
            events.append(("DataLoader", args, kwargs))
            self.config = PlanningConfig(initial_date=datetime(2025, 1, 1), forecast_months=2)
            self.forecast_actuals_months = 1
            self.periods = self.config.get_periods()
            self.materials = {
                "MAT-1": Material("MAT-1", "Material 1", ProductType.BULK_PRODUCT, "Family"),
                "SS-1": Material("SS-1", "Safety", ProductType.BULK_PRODUCT, "Family"),
            }
            self.bom = []
            self.safety_stock = {"SS-1": object()}
            self.machines = {}
            self.machine_groups = {}
            self.valuation_params = None
            self.purchased_and_produced = {}

        def load_all(self):
            events.append(("load_all",))
            return self

        def get_all_routings(self, material):
            return []

    class FakeForecastEngine:
        def __init__(self, data, actuals_months, forecast_months):
            events.append(("ForecastEngine", actuals_months, forecast_months))

        def calculate(self):
            events.append(("forecast.calculate",))
            return [_row("MAT-1", LineType.DEMAND_FORECAST.value)]

        def get_all_forecasts(self):
            return {"MAT-1": {"2025-01": 1.0}, "SS-1": {"2025-01": 2.0}}

    class FakeBOMEngine:
        def __init__(self, data):
            events.append(("BOMEngine",))

        def get_max_level(self):
            return 0

        def get_materials_at_level(self, level):
            return ["MAT-1"]

        def compute_dependent_requirements(self, mat_num, prod_plan):
            events.append(("compute_dependent_requirements", mat_num))
            return {}

    class FakeInventoryEngine:
        def __init__(self, data):
            events.append(("InventoryEngine",))

        def calculate_for_material(self, mat_num, forecast, dep_agg, dep_by_parent):
            events.append(("inventory.calculate", mat_num))
            return {
                "rows": [
                    _row(mat_num, LineType.TOTAL_DEMAND.value),
                    _row(mat_num, LineType.PRODUCTION_PLAN.value),
                    _row(mat_num, LineType.INVENTORY.value),
                ],
                "production_plan": {"2025-01": 1.0, "2025-02": 0.0},
                "purchase_receipt": None,
                "total_demand": {"2025-01": 1.0, "2025-02": 0.0},
                "purch_raw_need": {},
            }

    class FakeCapacityEngine:
        def __init__(self, data, production_plans, all_line_data=None):
            events.append(("CapacityEngine", dict(production_plans), bool(all_line_data)))

        def calculate(self):
            return {
                LineType.CAPACITY_UTILIZATION.value: [
                    _row("Z_M1", LineType.CAPACITY_UTILIZATION.value)
                ],
                LineType.FTE_REQUIREMENTS.value: [
                    _row("ZZ_GROUP", LineType.FTE_REQUIREMENTS.value)
                ],
            }

    class FakeValuePlanningEngine:
        def __init__(self, data, results):
            events.append(("ValuePlanningEngine", bool(results)))

        def calculate(self):
            return {
                LineType.CONSOLIDATION.value: [
                    _row("ZZZZZZ_TURNOVER", LineType.CONSOLIDATION.value)
                ]
            }

    monkeypatch.setattr(planning_module, "DataLoader", FakeDataLoader)
    monkeypatch.setattr(planning_module, "ForecastEngine", FakeForecastEngine)
    monkeypatch.setattr(planning_module, "BOMEngine", FakeBOMEngine)
    monkeypatch.setattr(planning_module, "InventoryEngine", FakeInventoryEngine)
    monkeypatch.setattr(planning_module, "CapacityEngine", FakeCapacityEngine)
    monkeypatch.setattr(planning_module, "ValuePlanningEngine", FakeValuePlanningEngine)

    engine = PlanningEngine(
        "synthetic.xlsm",
        planning_month="2025-02",
        months_actuals=1,
        months_forecast=2,
        config_overrides={"site": "NLX1"},
    )
    result = engine.run()

    assert result is engine
    assert engine.data.periods == ["2025-02", "2025-03"]
    assert set(engine.all_production_plans) == {"MAT-1", "SS-1"}
    assert engine.results[LineType.CAPACITY_UTILIZATION.value][0].material_number == "Z_M1"
    assert engine.value_results[LineType.CONSOLIDATION.value][0].material_number == "ZZZZZZ_TURNOVER"
    assert engine.summary["total_rows"] > 0
    assert [event[0] for event in events].index("ForecastEngine") < [event[0] for event in events].index("CapacityEngine")
