"""Fase 2.1 — comment (annotation) routes + session persistence."""

from types import SimpleNamespace

import pytest
from flask import Flask

from ui.routes.comments import create_comments_blueprint, comment_key

pytestmark = pytest.mark.no_fixture


@pytest.fixture
def comments_app():
    sess = {"id": "s1", "comments": {}}
    save_calls = []

    def get_active():
        return sess, None

    app = Flask(__name__)
    app.config["TESTING"] = True
    app.register_blueprint(create_comments_blueprint(
        get_active, lambda: save_calls.append(1)))
    return SimpleNamespace(app=app, client=app.test_client(), sess=sess, save_calls=save_calls)


def test_upsert_and_list_comment(comments_app):
    resp = comments_app.client.post("/api/comments", json={
        "scope": "01. Demand forecast", "target": "MAT-1", "period": "2026-01",
        "text": "Check met sales", "user": "sacha",
    })
    assert resp.status_code == 200
    key = comment_key("01. Demand forecast", "MAT-1", "2026-01")
    assert comments_app.sess["comments"][key]["text"] == "Check met sales"
    assert comments_app.sess["comments"][key]["user"] == "sacha"
    assert "updated_at" in comments_app.sess["comments"][key]
    assert comments_app.save_calls  # persisted

    listing = comments_app.client.get("/api/comments").get_json()
    assert key in listing["comments"]


def test_empty_text_deletes_comment(comments_app):
    comments_app.client.post("/api/comments", json={
        "scope": "machine", "target": "PBA01", "text": "let op"})
    comments_app.client.post("/api/comments", json={
        "scope": "machine", "target": "PBA01", "text": "   "})
    assert comments_app.sess["comments"] == {}


def test_delete_route_removes_comment(comments_app):
    comments_app.client.post("/api/comments", json={
        "scope": "machine", "target": "PBA01", "text": "x"})
    key = comment_key("machine", "PBA01", "")
    resp = comments_app.client.post("/api/comments/delete", json={"key": key})
    assert resp.get_json()["deleted"] is True
    assert comments_app.sess["comments"] == {}


def test_missing_scope_or_target_rejected(comments_app):
    assert comments_app.client.post("/api/comments", json={"text": "x"}).status_code == 400


def test_comments_survive_session_save_load(tmp_path):
    from ui.session_store import load_sessions_from_disk, save_sessions_to_disk

    sessions = {"s1": {"id": "s1", "engine": None, "parameters": {"planning_month": "2025-12"},
                       "comments": {comment_key("machine", "PBA01", ""): {
                           "text": "hoi", "user": "u", "updated_at": "2026-07-11T00:00:00"}}}}
    store = tmp_path / "sessions_store.json"
    save_sessions_to_disk(sessions, "s1", store, lambda s, e: {})
    loaded, _ = load_sessions_from_disk(store)
    assert loaded["s1"]["comments"][comment_key("machine", "PBA01", "")]["text"] == "hoi"


def test_export_comments_sheet(tmp_path):
    import openpyxl
    from ui.routes.exports import _append_comments_sheet

    wb = openpyxl.Workbook()
    wb.active.title = "Planning sheet"
    path = tmp_path / "out.xlsx"
    wb.save(path)

    _append_comments_sheet(str(path), {
        comment_key("01. Demand forecast", "MAT-1", "2026-01"): {
            "text": "Afstemmen met productie", "user": "abdel", "updated_at": "2026-07-11T01:00:00"},
    })
    wb2 = openpyxl.load_workbook(path)
    assert "Opmerkingen" in wb2.sheetnames
    ws = wb2["Opmerkingen"]
    assert ws.cell(2, 4).value == "Afstemmen met productie"
    assert ws.cell(2, 5).value == "abdel"


def test_export_comments_sheet_survives_control_characters(tmp_path):
    """F15: control chars in comment text must not fail the export."""
    import openpyxl
    from ui.routes.exports import _append_comments_sheet

    wb = openpyxl.Workbook()
    wb.active.title = "Planning sheet"
    path = tmp_path / "out.xlsx"
    wb.save(path)

    _append_comments_sheet(str(path), {
        comment_key("machine", "PBA01", ""): {
            "text": "regel1\x0bregel2\x00einde", "user": "u\x0c", "updated_at": "2026-07-11"},
    })
    ws = openpyxl.load_workbook(path)["Opmerkingen"]
    assert ws.cell(2, 4).value == "regel1regel2einde"
