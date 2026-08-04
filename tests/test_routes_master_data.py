"""Masterdata-routes + werkboek-vrije upload/calculate-flow."""

import json
from types import SimpleNamespace

import pytest
from flask import Flask

from modules.master_data import serialize_master
from tests.master_fixtures import fake_master_loader, write_extract_files
from ui import master_store
from ui.routes.master_data import create_master_data_blueprint

pytestmark = pytest.mark.no_fixture


@pytest.fixture
def store_env(tmp_path):
    """Eigen store-pad per test; module-global netjes herstellen."""
    previous = master_store.get_store_path()
    path = tmp_path / 'master_store.json'
    master_store.set_store_path(path)
    yield SimpleNamespace(path=path, tmp=tmp_path)
    master_store.set_store_path(previous) if previous else master_store.set_store_path(path)


def _seed_store(store_env):
    master = json.loads(json.dumps(serialize_master(fake_master_loader()), default=str))
    master_store.save_master_store(store_env.path, master, source_filename='seed.xlsm')
    master_store.set_store_path(store_env.path)  # cache verversen
    return master


@pytest.fixture
def md_app(store_env):
    app = Flask(__name__)
    app.config['TESTING'] = True
    app.register_blueprint(create_master_data_blueprint(
        {}, lambda: store_env.tmp / 'uploads'))
    return SimpleNamespace(app=app, client=app.test_client(), store=store_env)


def test_status_reports_absence_and_presence(md_app):
    assert md_app.client.get('/api/master_data').get_json() == {'exists': False}
    _seed_store(md_app.store)
    body = md_app.client.get('/api/master_data').get_json()
    assert body['exists'] and body['version'] == 1
    assert body['counts']['materials'] == 2 and body['counts']['machines'] == 1


def test_get_dataset_and_unknown(md_app):
    assert md_app.client.get('/api/master_data/materials').status_code == 400  # geen store
    _seed_store(md_app.store)
    body = md_app.client.get('/api/master_data/materials').get_json()
    assert [m['material_number'] for m in body['value']] == ['M1', 'M2']
    assert md_app.client.get('/api/master_data/nonsense').status_code == 404


def test_patch_validates_by_hydration_and_bumps_version(md_app):
    master = _seed_store(md_app.store)

    # Geldige bewerking: naam aanpassen.
    materials = [dict(m) for m in master['materials']]
    materials[0]['name'] = 'Parent (hernoemd)'
    resp = md_app.client.patch('/api/master_data/materials',
                               json={'value': materials})
    assert resp.status_code == 200
    body = resp.get_json()
    assert body['version'] == 2 and body['requires_recalculate'] is True
    stored = master_store.load_master_store(md_app.store.path)
    assert stored['master']['materials'][0]['name'] == 'Parent (hernoemd)'
    assert stored['edited_at'] and stored['source_filename'] == 'seed.xlsm'

    # Verkeerd type → 400; hydratie-fout (kapot record) → 400 met reden.
    assert md_app.client.patch('/api/master_data/materials',
                               json={'value': {'nope': 1}}).status_code == 400
    broken = [dict(materials[0])]
    broken[0].pop('product_type')
    resp = md_app.client.patch('/api/master_data/materials', json={'value': broken})
    assert resp.status_code == 400
    assert 'Wijziging geweigerd' in resp.get_json()['error']
    # Mislukte patch mag de versie niet ophogen.
    assert master_store.load_master_store(md_app.store.path)['version'] == 2


def test_import_requires_source_and_confirm_flow(md_app, tmp_path):
    # Geen upload en geen geconfigureerd master_file → 400.
    resp = md_app.client.post('/api/master_data/import', data={})
    assert resp.status_code == 400
    assert 'Geen bronbestand' in resp.get_json()['error']


def test_workbook_free_upload_and_calculate(flask_test_app, store_env, tmp_path):
    """Multi-upload zonder basisbestand: master-store + 4 extracts → sessie
    zonder file_path; /api/calculate rekent vanuit de store."""
    _seed_store(store_env)
    extracts = write_extract_files(tmp_path)

    files = {}
    for form_key, dict_key in [('bom_file', 'bom'), ('routing_file', 'routing'),
                               ('stock_file', 'stock'), ('forecast_file', 'forecast')]:
        files[form_key] = (open(extracts[dict_key], 'rb'),
                          f'{dict_key}_extract.xlsx')
    resp = flask_test_app.client.post(
        '/api/upload', data={**files, 'custom_name': 'Werkboek-vrij'},
        content_type='multipart/form-data')
    body = resp.get_json()
    assert resp.status_code == 200 and body.get('success'), body
    session_id = body['session_id']
    sess = flask_test_app.sessions[session_id]
    assert sess['file_path'] == ''  # geen basiswerkboek
    assert set(sess['extract_files']) == {'bom', 'routing', 'stock', 'forecast'}

    calc = flask_test_app.client.post('/api/calculate', json={
        'planning_month': None, 'months_actuals': 1, 'months_forecast': 3})
    calc_body = calc.get_json()
    assert calc.status_code == 200 and calc_body.get('success'), calc_body
    engine = sess['engine']
    assert engine is not None
    assert engine.data.periods == ['2026-01', '2026-02', '2026-03']
    assert 'M1' in engine.data.materials and 'PBA01' in engine.data.machines


def test_workbook_free_session_rebuild_uses_latest_store(store_env, tmp_path):
    """Herstart-/rebuild-pad: build_clean_engine_for_session bouwt een sessie
    zonder file_path vanuit de LAATSTE master-store; zonder store faalt hij
    netjes (None) in plaats van te crashen."""
    from ui.engine_rebuild import build_clean_engine_for_session

    extracts = write_extract_files(tmp_path)
    sess = {
        'id': 'wf', 'file_path': '', 'extract_files': extracts,
        'parameters': {'planning_month': None, 'months_actuals': 1,
                       'months_forecast': 3},
        'pending_edits': {}, 'value_aux_overrides': {}, 'machine_overrides': {},
        'inventory_overrides': {}, 'capacity_overrides': {}, 'engine': None,
    }
    # Zonder store: nette weigering.
    assert build_clean_engine_for_session(sess, {}) is None

    _seed_store(store_env)
    import contextlib
    import io as _io
    with contextlib.redirect_stdout(_io.StringIO()):
        engine = build_clean_engine_for_session(sess, {})
    assert engine is not None
    assert engine.data.periods == ['2026-01', '2026-02', '2026-03']
    assert 'M1' in engine.data.materials


# ------------------------------------------------------- masterwerkboek-routes


def _export_workbook(md_app):
    res = md_app.client.get('/api/master_workbook/export')
    assert res.status_code == 200
    path = md_app.store.tmp / 'export.xlsx'
    path.write_bytes(res.data)
    return path


def test_workbook_export_then_import_is_a_noop_diff(md_app):
    _seed_store(md_app.store)
    path = _export_workbook(md_app)

    with open(path, 'rb') as f:
        res = md_app.client.post('/api/master_workbook/import',
                                 data={'file': (f, 'SOP_Masterdata_NLX1.xlsx')},
                                 content_type='multipart/form-data')
    body = res.get_json()
    assert body['needs_confirm'] is True
    assert body['stale_export'] is None  # zelfde versie
    d = body['diff']
    assert d['materials_added'] == [] and d['materials_removed'] == []
    assert d['materials_changed'] == [] and d['datasets_changed'] == []


def test_workbook_import_applies_edits_and_deactivates_missing(md_app):
    _seed_store(md_app.store)
    path = _export_workbook(md_app)

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb['Materialen']
    headers = [c.value for c in ws[1]]
    # M1 hernoemen, M2-rij verwijderen (→ moet gedeactiveerd worden, niet weg)
    ws.cell(row=2, column=headers.index('name') + 1, value='HERNOEMD')
    assert ws.cell(row=3, column=headers.index('material_number') + 1).value == 'M2'
    ws.delete_rows(3)
    wb.save(str(path))

    with open(path, 'rb') as f:
        res = md_app.client.post('/api/master_workbook/import',
                                 data={'file': (f, 'wb.xlsx'), 'confirm': 'true'},
                                 content_type='multipart/form-data')
    body = res.get_json()
    assert body['success'] is True
    assert body['deactivated'] == ['M2']

    stored = master_store.get_current_master_record()['master']
    by_num = {m['material_number']: m for m in stored['materials']}
    assert by_num['M1']['name'] == 'HERNOEMD'
    assert by_num['M2']['is_active'] is False        # gedeactiveerd, niet verwijderd
    assert by_num['M2']['name'] == 'Child'           # overige velden behouden


def test_workbook_import_refuses_other_site(md_app):
    _seed_store(md_app.store)  # site NLX1
    path = _export_workbook(md_app)

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb['_Meta']
    for row in ws.iter_rows():
        if row[0].value == 'site':
            row[1].value = 'NLK1'
    wb.save(str(path))

    with open(path, 'rb') as f:
        res = md_app.client.post('/api/master_workbook/import',
                                 data={'file': (f, 'wb.xlsx'), 'confirm': 'true'},
                                 content_type='multipart/form-data')
    assert res.status_code == 400
    assert 'NLK1' in res.get_json()['error']


def test_workbook_import_flags_stale_export(md_app):
    _seed_store(md_app.store)
    path = _export_workbook(md_app)
    # Store muteren ná de export: versie loopt op.
    record = master_store.get_current_master_record()
    master_store.save_master_store(md_app.store.path, record['master'],
                                   previous=record, edited=True)

    with open(path, 'rb') as f:
        res = md_app.client.post('/api/master_workbook/import',
                                 data={'file': (f, 'wb.xlsx')},
                                 content_type='multipart/form-data')
    body = res.get_json()
    assert body['needs_confirm'] is True
    assert body['stale_export'] == {'exported_from_version': 1, 'store_version': 2}


def test_workbook_import_preserves_store_actuals(md_app):
    master = _seed_store(md_app.store)
    master['purchase']['actuals'] = {'M2': {'2026-01': 7.0}}
    record = master_store.get_current_master_record()
    master_store.save_master_store(md_app.store.path, master, previous=record)
    path = _export_workbook(md_app)

    with open(path, 'rb') as f:
        res = md_app.client.post('/api/master_workbook/import',
                                 data={'file': (f, 'wb.xlsx'), 'confirm': 'true'},
                                 content_type='multipart/form-data')
    assert res.get_json()['success'] is True
    stored = master_store.get_current_master_record()['master']
    assert stored['purchase']['actuals'] == {'M2': {'2026-01': 7.0}}


def test_workbook_export_without_store_is_400(md_app):
    assert md_app.client.get('/api/master_workbook/export').status_code == 400


def test_patch_refreshes_mirror(md_app, tmp_path):
    from ui import master_mirror
    _seed_store(md_app.store)
    body = md_app.client.get('/api/master_data/materials').get_json()
    value = body['value']
    value[0]['name'] = 'VIA GRID'
    res = md_app.client.patch('/api/master_data/materials', json={'value': value})
    assert res.get_json()['success'] is True
    status = master_mirror.mirror_status()
    assert status['stale'] is False
    assert status['path'].endswith('SOP_Masterdata_NLX1.xlsx')
    import openpyxl
    wb = openpyxl.load_workbook(status['path'], read_only=True)
    ws = wb['Materialen']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    assert any('VIA GRID' in [str(v) for v in row] for row in rows[1:])


def test_add_missing_material_and_reactivate(md_app):
    master = _seed_store(md_app.store)

    # Nieuw materiaal (consistentiebanner-actie)
    res = md_app.client.post('/api/master_data/materials/add',
                             json={'material': 'M9', 'name': 'Uit extract'})
    body = res.get_json()
    assert body['success'] and body['action'] == 'added'
    stored = master_store.get_current_master_record()['master']
    by_num = {m['material_number']: m for m in stored['materials']}
    assert by_num['M9']['is_active'] is True
    assert by_num['M9']['product_type'] == 'Other'

    # Gedeactiveerd exemplaar wordt geheractiveerd, niet gedupliceerd
    value = [dict(m) for m in stored['materials']]
    next(m for m in value if m['material_number'] == 'M2')['is_active'] = False
    md_app.client.patch('/api/master_data/materials', json={'value': value})
    res = md_app.client.post('/api/master_data/materials/add', json={'material': 'M2'})
    assert res.get_json()['action'] == 'reactivated'
    stored = master_store.get_current_master_record()['master']
    assert sum(1 for m in stored['materials'] if m['material_number'] == 'M2') == 1

    # Al actief → idempotent succes (promote-flow / dubbelklik op de
    # banner-knop mag geen rode foutmelding geven)
    res = md_app.client.post('/api/master_data/materials/add', json={'material': 'M1'})
    body = res.get_json()
    assert res.status_code == 200 and body['action'] == 'already_active'
    assert body['requires_recalculate'] is False


def test_full_product_wizard_add_lands_in_all_datasets(md_app):
    """Masterdata-productwizard: één POST zet het product atomair in alle
    relevante datasets (materiaal, veiligheidsvoorraad, inkoop, prijs,
    kost) — de tabellen zijn de enige bron van waarheid."""
    _seed_store(md_app.store)
    res = md_app.client.post('/api/master_data/materials/add', json={
        'material': 'M9', 'name': 'Wizardproduct',
        'product_type': 'Raw Material', 'product_family': 'FAM',
        'spc_product': 'SPC9', 'default_inventory_value': 12.5,
        'safety_stock': {'safety_stock': 40, 'lot_size': 10},
        'purchase': {'lead_time': 2, 'moq': 25},
        'sales_price': {'price_per_unit': 80, 'volume': 1200},
        'material_cost': {'cost_per_unit': 55.5},
    })
    body = res.get_json()
    assert body['success'] and body['action'] == 'added', body

    m = master_store.get_current_master_record()['master']
    mat = next(x for x in m['materials'] if x['material_number'] == 'M9')
    assert mat['product_type'] == 'Raw Material'
    assert mat['spc_product'] == 'SPC9'
    assert mat['default_inventory_value'] == 12.5
    assert m['safety_stock']['M9']['safety_stock'] == 40.0
    assert m['safety_stock']['M9']['lot_size'] == 10.0
    assert m['purchase']['lead_times']['M9'] == 2
    assert m['purchase']['moq']['M9'] == 25.0
    assert 'M9' in m['purchase']['sheet_materials']
    assert m['sales_prices']['M9']['ex_works_revenue'] == 80 * 1200
    assert m['sales_prices']['M9']['volume_2025'] == 1200.0
    assert m['material_costs']['M9']['cost_per_unit'] == 55.5
    assert m['material_costs']['M9']['plant_code'] == m['config']['site']

    # Ongeldig getal → nette 400, niets gemuteerd
    v_before = master_store.get_current_master_record()['version']
    res = md_app.client.post('/api/master_data/materials/add', json={
        'material': 'M10', 'name': 'X', 'purchase': {'lead_time': 'abc'}})
    assert res.status_code == 400
    assert master_store.get_current_master_record()['version'] == v_before

    # Secties op een al-actief materiaal → update, geen already_active
    res = md_app.client.post('/api/master_data/materials/add', json={
        'material': 'M9', 'safety_stock': {'safety_stock': 60, 'lot_size': 10}})
    assert res.get_json()['action'] == 'updated'
    m = master_store.get_current_master_record()['master']
    assert m['safety_stock']['M9']['safety_stock'] == 60.0

