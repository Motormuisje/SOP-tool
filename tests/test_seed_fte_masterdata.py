"""Seed-script voor de F2-CF masterdata (tools/seed_fte_masterdata.py).

Draait op een synthetisch werkboek met dezelfde vorm als het klantmodel — het
echte bestand bevat klantdata en staat niet in de repo. Wat hier bewaakt wordt
is de LEESLOGICA: de bruto→netto FTE-afleiding, de draaitabel-doorvulling op
het PEER-blad, en de vertaaltabel klantnaam → SAP-code.
"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'tools'))

from modules.models import FTE_PARAM_DEFAULTS  # noqa: E402
from seed_fte_masterdata import (  # noqa: E402
    EXPECTED_SITE,
    MODEL_TO_SAP,
    SeedError,
    _merge_dataset,
    _merge_machines,
    apply_seed,
    read_fte,
    read_maintenance,
    read_model_machines,
    read_peer,
    read_truck_rows,
    verify_mapping,
)

pytestmark = pytest.mark.no_fixture


@pytest.fixture
def workbook():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fte = wb.create_sheet('FTE')
    for cell, value in (
        ('B5', 'Effective days'), ('D5', 260), ('E5', 8), ('F5', 2080),
        ('B6', 'Holiday'), ('D6', 28), ('E6', 8), ('F6', -224),
        ('B7', 'ATV'), ('D7', 14), ('E7', 8), ('F7', -112),
        ('B8', 'Public holidays in workweek'), ('D8', 6), ('E8', 8), ('F8', -48),
        ('B9', 'Working hours after holidays'), ('F9', 1696),
        ('B10', 'Sick leave'), ('C10', 0.1), ('F10', -169.6),
        ('B11', 'Training'), ('C11', 0.02), ('F11', -33.92),
        ('B13', 'Working hours per FTE per yr'), ('F13', 1492.48),
        ('B16', 'Bezettingsgraad'), ('D16', 0.85),
        ('B17', '2-ploegen'), ('C17', 4160),
        ('B18', '3-ploegen'), ('C18', 6240),
        ('B19', '24/7'), ('C19', 8760),
    ):
        fte[cell] = value

    model = wb.create_sheet('OEE Model MST ')
    # Kolomkoppen: rij 7 = OEE, rij 8 = capaciteit, rij 10 = naam.
    for column, (name, oee, capacity) in enumerate(
            [('MRL', 0.7, 50), ('PM01', 0.8, 3.5), ('Pneu', 0.7, 35)], start=8):
        model.cell(row=10, column=column, value=name)
        model.cell(row=7, column=column, value=oee)
        model.cell(row=8, column=column, value=capacity)
    # Trucks
    for cell, value in (
        ('D185', 'Container'), ('C185', 0.35), ('E185', 1.2), ('F185', 14644.6),
        ('G185', 22), ('H185', 665.663), ('I185', 0.535213),
        ('D186', 'Pallet/BB'), ('E186', 1), ('F186', 27197.1), ('G186', 24),
        ('H186', 1133.21), ('I186', 0.759281),
        ('D192', 'Containers'), ('E192', 1.5), ('F192', 7035.22), ('G192', 22),
        ('H192', 319.783), ('I192', 0.321394),
        ('B198', 'FTE Direct Maintenance'), ('D198', 9), ('E198', 'x / FTE'),
        ('D200', 67134.6), ('E200', 7459.4),
        ('D201', 1750000), ('E201', 194444),
    ):
        model[cell] = value

    peer = wb.create_sheet('PEER_Capacity')
    peer.append(['Facility Description', 'FullDescription', 'FullDescription',
                 'Throughput (tonnes/hour)'])
    # Draaitabelvorm: de installatienaam staat alleen op de eerste regel.
    peer.append(['NL-PP-Maastricht', 'AMB Ball Mill 3 (PML06)', 'Total', 27.79])
    peer.append([None, None, 'Ca-Carb A125 PO (600003728)', 26.23])
    peer.append([None, None, 'Dolomite DS19 PO (600003818)', 38.6])
    peer.append([None, 'AMB Bagging Machine 7 (PBA07)', 'Total', 6.21])
    peer.append([None, None, 'BAGGING', 6.21])
    return wb


def test_fte_derivation_matches_the_workbook(workbook):
    params, effective, shift_hours, checks = read_fte(workbook['FTE'])

    assert params == {
        'utilization_rate': 0.85, 'gross_hours_per_year': 2080.0,
        'leave_hours_per_year': 224.0, 'adv_hours_per_year': 112.0,
        'holiday_hours_per_year': 48.0, 'illness_pct': 0.1, 'training_pct': 0.02,
    }
    assert effective == pytest.approx(1492.48)
    assert shift_hours['3-shift system'] == pytest.approx(520.0)
    assert all('AFWIJKING' not in check for check in checks[:2])
    # De derde controleregel toont juist het verschil met stapelend rekenen.
    assert '+3.39' in checks[2]


def test_peer_forward_fills_the_pivot(workbook):
    """Zonder doorvullen leest élke productregel als 'geen machine' en valt de
    hele tabel weg — dat was de bug die 71 benchmarks tot 11 reduceerde."""
    rows, skipped = read_peer(workbook['PEER_Capacity'])
    by_key = {(r['machine_code'], r['material_number']): r['throughput'] for r in rows}

    assert by_key[('PML06', '')] == pytest.approx(27.79)          # installatietotaal
    assert by_key[('PML06', '600003728')] == pytest.approx(26.23)  # productregel
    assert by_key[('PML06', '600003818')] == pytest.approx(38.6)
    assert by_key[('PBA07', '')] == pytest.approx(6.21)
    assert any('BAGGING' in s for s in skipped)                    # geen materiaalnummer


def test_truck_rows_are_read_with_their_derivation(workbook):
    loading, unloading = read_truck_rows(workbook['OEE Model MST '])

    container = next(r for r in loading if r['label'] == 'Container')
    assert container['tons_per_truck'] == 22
    assert container['hours_per_truck'] == pytest.approx(1.2)
    assert container['volume'] / container['tons_per_truck'] == pytest.approx(
        container['trucks'], rel=1e-3)
    assert container['trucks'] * container['hours_per_truck'] / 1492.48 == pytest.approx(
        container['fte'], rel=1e-3)
    assert [r['label'] for r in unloading] == ['Containers']


def test_maintenance_is_a_headcount_not_a_ratio(workbook):
    """Rij 198 leest als '9 x / FTE'. Het klantmodel deelt de kengetallen in
    rij 200/201 dóór 9, dus 9 is het AANTAL onderhouds-FTE — niet '9 machines
    per FTE', zoals het plan aanvankelijk aannam."""
    maintenance = read_maintenance(workbook['OEE Model MST '])

    assert maintenance['fte'] == 9
    assert maintenance['machine_hours_total'] / 9 == pytest.approx(
        maintenance['machine_hours_per_fte'], rel=1e-4)
    assert maintenance['opex_total'] / 9 == pytest.approx(
        maintenance['opex_per_fte'], rel=1e-4)


def test_mapping_is_verified_against_the_masterdata_oee(workbook):
    model_machines = read_model_machines(workbook['OEE Model MST '])
    assert set(model_machines) == {'MRL', 'PM01', 'Pneu'}

    sap = {'PSS13': {'oee': 0.7}, 'PML01': {'oee': 0.8}, 'PPM09': {'oee': 0.8}}
    table = verify_mapping(model_machines, sap)
    joined = '\n'.join(table)

    assert '| MRL | PSS13 | 0.7 | 0.7 | 50 | OK |' in joined
    assert '| PM01 | PML01 | 0.8 | 0.8 | 3.5 | OK |' in joined
    # Pneu: model 0,7 vs masterdata 0,8 — moet als afwijking naar boven komen
    # in plaats van stilzwijgend te worden gekoppeld.
    assert 'Pneu' in joined and 'OEE WIJKT AF' in joined


class TestRobustAgainstAShiftedSheet:
    """Vaste celadressen braken zodra de klant één rij invoegde: F13 wees dan
    naar een lege cel (uren per FTE 0, door hydratie stil 1492) en de
    ploegurenrijen schoven op, waardoor elke 3-ploegenmachine 347 in plaats
    van 520 beschikbare uren kreeg. Stil, en de controleregel meldde 'OK'."""

    def test_inserted_row_does_not_shift_the_values(self, workbook):
        fte = workbook['FTE']
        fte.insert_rows(4)          # extra kop boven 'Effective days'

        params, effective, shift_hours, _ = read_fte(fte)

        assert effective == pytest.approx(1492.48)
        assert params['gross_hours_per_year'] == 2080.0
        assert shift_hours['3-shift system'] == pytest.approx(520.0)

    def test_missing_label_aborts_instead_of_seeding_zeros(self, workbook):
        fte = workbook['FTE']
        fte['B13'] = 'Iets anders'   # 'Working hours per FTE per yr' weg

        with pytest.raises(SeedError, match='Working hours per FTE per yr'):
            read_fte(fte)

    def test_broken_derivation_aborts(self, workbook):
        fte = workbook['FTE']
        fte['F13'] = 1200            # past niet bij 1696 - 10% - 2%

        with pytest.raises(SeedError, match='afleiding'):
            read_fte(fte)

    def test_a_duplicate_label_aborts_instead_of_reading_the_first(self, workbook):
        """Een klant zet een overzichtsregel met dezelfde tekst boven het blok.
        Voor de bruto->netto-velden vangen de afleidingscontroles dat af, maar
        voor de bezettingsgraad en de ploeguren merkte niets het — die werden
        dan stil uit de verkeerde rij gelezen."""
        fte = workbook['FTE']
        fte['B2'] = 'Bezettingsgraad'
        fte['D2'] = 0.5

        with pytest.raises(SeedError, match='meerdere rijen'):
            read_fte(fte)

    def test_implausible_shift_hours_abort(self, workbook):
        """347 jaaruren is geen ploegensysteem maar een verschoven cel."""
        fte = workbook['FTE']
        fte['C18'] = 347

        with pytest.raises(SeedError, match='onmogelijk'):
            read_fte(fte)

    def test_impossible_occupancy_aborts(self, workbook):
        fte = workbook['FTE']
        fte['D16'] = 1.7

        with pytest.raises(SeedError, match='bezettingsgraad'):
            read_fte(fte)


class TestSecondRunIsSafe:
    """De seed zet records neer die de klant daarna bijstelt (een activiteit
    aanzetten, een norm corrigeren). Een tweede --apply mag dat niet stil
    terugdraaien — dat maakte het script gevaarlijker dan nuttig."""

    def test_existing_records_are_left_alone_by_default(self):
        existing = {'A': {'operators_per_hour': 2.0}}   # klant corrigeerde dit
        seeded = {'A': {'operators_per_hour': 1.0}, 'B': {'operators_per_hour': 1.0}}

        merged, stats = _merge_dataset(existing, seeded, overwrite=False)

        assert merged['A']['operators_per_hour'] == 2.0
        assert merged['B']['operators_per_hour'] == 1.0
        assert stats['toegevoegd'] == ['B']
        assert stats['overschreven'] == []

    def test_overwrite_flag_replaces_and_reports_it(self):
        existing = {'A': {'operators_per_hour': 2.0}}
        seeded = {'A': {'operators_per_hour': 1.0}}

        merged, stats = _merge_dataset(existing, seeded, overwrite=True)

        assert merged['A']['operators_per_hour'] == 1.0
        assert stats['overschreven'] == ['A']

    def test_machine_edits_in_the_app_survive(self):
        """OEE schaalt de capaciteit rechtstreeks; terugzetten naar de
        werkboekwaarde verschuift Line 09-12 en de hele werkbank."""
        existing = [{'machine_code': 'PML06', 'oee': 0.99, 'name': 'Handmatig gecorrigeerd',
                     'availability_by_period': {'2025-01': 0.9}}]
        seeded = [{'machine_code': 'PML06', 'oee': 0.70, 'name': 'Ball Mill 3',
                   'availability_by_period': {}},
                  {'machine_code': 'PML05', 'oee': 0.70, 'name': 'Ball Mill 2',
                   'availability_by_period': {}}]

        merged, notes = _merge_machines(existing, seeded, overwrite=False)
        by_code = {m['machine_code']: m for m in merged}

        assert by_code['PML06']['oee'] == 0.99
        assert by_code['PML06']['name'] == 'Handmatig gecorrigeerd'
        assert by_code['PML05']['oee'] == 0.70            # nieuw, dus toegevoegd
        assert any('PML06' in n and 'ongewijzigd' in n for n in notes)

    def test_month_data_of_the_store_wins_on_overwrite(self):
        existing = [{'machine_code': 'PML06', 'oee': 0.99,
                     'availability_by_period': {'2025-01': 0.9}}]
        seeded = [{'machine_code': 'PML06', 'oee': 0.70, 'availability_by_period': {}}]

        merged, _ = _merge_machines(existing, seeded, overwrite=True)

        assert merged[0]['oee'] == 0.70                                   # overschreven
        assert merged[0]['availability_by_period'] == {'2025-01': 0.9}    # maanddata blijft


class TestFirstApplyActuallySeeds:
    """serialize_master schrijft ALTIJD een params-blok (gevuld met defaults),
    dus 'er staan al parameters' was nooit onwaar. De allereerste --apply
    seedde de FTE-parameters en de uren per FTE daardoor nooit — precies het
    getal waar de hele seed om begon."""

    def _store(self, tmp_path, params=None, hours=1492.0):
        from ui import master_store

        path = tmp_path / 'master_store.json'
        master_store.set_store_path(path)
        fte = {'fte_hours_per_year': hours, 'shift_hours': {}}
        if params is not None:
            fte['params'] = params
        master_store.save_master_store(path, {'config': {'site': 'NLX1'}, 'fte': fte},
                                       source_filename='seed.xlsm')
        master_store.set_store_path(path)
        return path

    def _datasets(self):
        return {'fte_params': {'utilization_rate': 0.85, 'gross_hours_per_year': 2080.0,
                               'leave_hours_per_year': 224.0, 'adv_hours_per_year': 112.0,
                               'holiday_hours_per_year': 48.0, 'illness_pct': 0.1,
                               'training_pct': 0.02},
                'fte_hours_per_year': 1492.48,
                'shift_hours': {'3-shift system': 520.0},
                'staffing_norms': {}, 'indirect_activities': {},
                'benchmark_throughput': {}}

    def test_defaults_in_the_store_do_not_block_the_seed(self, tmp_path):
        from ui import master_store

        previous = master_store.get_store_path()
        path = self._store(tmp_path, params=dict(FTE_PARAM_DEFAULTS))
        try:
            master, changes = apply_seed(path, self._datasets(), None,
                                         verify_only=True, expected_site=EXPECTED_SITE)
            assert master['fte']['params']['gross_hours_per_year'] == 2080.0
            assert master['fte']['fte_hours_per_year'] == pytest.approx(1492.48)
            assert any('params' in c for c in changes)
        finally:
            master_store.set_store_path(previous) if previous else None

    def test_customer_edited_parameters_are_left_alone(self, tmp_path):
        from ui import master_store

        previous = master_store.get_store_path()
        edited = dict(FTE_PARAM_DEFAULTS)
        edited['utilization_rate'] = 0.9          # klant heeft dit aangepast
        path = self._store(tmp_path, params=edited)
        try:
            master, changes = apply_seed(path, self._datasets(), None,
                                         verify_only=True, expected_site=EXPECTED_SITE)
            assert master['fte']['params']['utilization_rate'] == 0.9
            assert any('ONGEWIJZIGD' in c for c in changes)
        finally:
            master_store.set_store_path(previous) if previous else None

    def test_shift_hours_changes_are_reported(self, tmp_path):
        """Ploeguren schalen élke machine; ze muteren zonder regel in het
        wijzigingsoverzicht is precies de stille wijziging die het rapport
        moet voorkomen."""
        from ui import master_store

        previous = master_store.get_store_path()
        path = self._store(tmp_path, params=dict(FTE_PARAM_DEFAULTS))
        try:
            _, changes = apply_seed(path, self._datasets(), None, verify_only=True,
                                    expected_site=EXPECTED_SITE)
            assert any('ploeguren' in c.lower() for c in changes)
        finally:
            master_store.set_store_path(previous) if previous else None

    def test_the_occupancy_rate_is_called_out_separately(self, tmp_path):
        """utilization_rate zit in het params-blok maar REKENT mee (staffed
        FTE). 'Verandert op zichzelf geen getal' klopte daar niet."""
        from ui import master_store

        previous = master_store.get_store_path()
        edited = dict(FTE_PARAM_DEFAULTS)
        edited['utilization_rate'] = 0.9
        path = self._store(tmp_path, params=edited)
        try:
            _, changes = apply_seed(path, self._datasets(), None, verify_only=True,
                                    overwrite=True, expected_site=EXPECTED_SITE)
            assert any('bezettingsgraad' in c.lower() and '0.9' in c for c in changes)
        finally:
            master_store.set_store_path(previous) if previous else None


class TestSiteGuard:
    """Met SOP_APP_DATA_DIR gezet wijst de default store naar een ZUSTERSITE.
    Maastricht-normen daar inschrijven is stille datavervuiling."""

    def test_a_foreign_site_store_is_refused(self, tmp_path):
        from ui import master_store

        previous = master_store.get_store_path()
        path = tmp_path / 'master_store.json'
        master_store.set_store_path(path)
        master_store.save_master_store(path, {'config': {'site': 'NLK1'}},
                                       source_filename='wsk')
        master_store.set_store_path(path)
        datasets = {'fte_params': {}, 'fte_hours_per_year': 1492.48, 'shift_hours': {},
                    'staffing_norms': {}, 'indirect_activities': {},
                    'benchmark_throughput': {}}
        try:
            with pytest.raises(SystemExit, match='NLK1'):
                apply_seed(path, datasets, None, verify_only=True,
                           expected_site=EXPECTED_SITE)
        finally:
            master_store.set_store_path(previous) if previous else None

    def test_force_site_allows_it(self, tmp_path):
        from ui import master_store

        previous = master_store.get_store_path()
        path = tmp_path / 'master_store.json'
        master_store.set_store_path(path)
        master_store.save_master_store(path, {'config': {'site': 'NLK1'}},
                                       source_filename='wsk')
        master_store.set_store_path(path)
        datasets = {'fte_params': {}, 'fte_hours_per_year': 1492.48, 'shift_hours': {},
                    'staffing_norms': {}, 'indirect_activities': {},
                    'benchmark_throughput': {}}
        try:
            _, changes = apply_seed(path, datasets, None, verify_only=True,
                                    expected_site='')
            assert changes
        finally:
            master_store.set_store_path(previous) if previous else None


def test_every_mapped_machine_has_a_unique_purpose():
    """Twee klantnamen mogen naar dezelfde SAP-machine wijzen (PE20 komt in het
    model zowel als big-bag- als als zaklijn voor), maar een SAP-code mag nooit
    leeg of onbekend zijn."""
    for name, (code, label) in MODEL_TO_SAP.items():
        assert code and code.isalnum(), f'{name} heeft geen bruikbare SAP-code'
        assert label, f'{name} heeft geen omschrijving'
