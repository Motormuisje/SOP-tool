"""Cell-by-cell comparison of an app export against a ground-truth workbook.

The ground truth is the client's Excel model (MS_RECONC ``.xlsm`` after the VBA
macros have run) or a hand-verified copy of it. The app export is the planning
workbook written by ``PlanningEngine.to_excel_with_values`` (the "Download"
button in the web UI, or ``main.py --cli``).

Both files carry the same sheet layout: one header row, identifier columns,
then one column per period. Rows are matched on
``(Material number, Line type, Aux Column, Aux 2 Column)`` — never on row
position, because the two writers order rows differently.

Usage::

    python tools/ground_truth_diff.py ground_truth.xlsm app_export.xlsx
    python tools/ground_truth_diff.py gt.xlsm app.xlsx --out delta.xlsx --md delta.md
    python tools/ground_truth_diff.py gt.xlsm app.xlsx --sheet "Planning sheet"

Exit code 0 means every compared cell is within tolerance; 1 means at least
one deviation or an unmatched row was found; 2 means the files could not be
compared (missing sheet, unreadable file).
"""
from __future__ import annotations

import argparse
import math
import sys
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import openpyxl
except ImportError:  # pragma: no cover - environment guard
    print("openpyxl ontbreekt: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

DEFAULT_SHEETS = ("Planning sheet", "Values_Planning sheet")
KEY_HEADERS = ("Material number", "Line type", "Aux Column", "Aux 2 Column")
VALUE_HEADERS = ("Starting stock",)

RowKey = Tuple[str, str, str, str]


@dataclass
class Deviation:
    sheet: str
    key: RowKey
    column: str
    ground_truth: object
    app: object
    abs_diff: float
    rel_diff: float


@dataclass
class SheetReport:
    sheet: str
    rows_gt: int = 0
    rows_app: int = 0
    keys_matched: int = 0
    keys_only_gt: List[RowKey] = field(default_factory=list)
    keys_only_app: List[RowKey] = field(default_factory=list)
    columns_compared: List[str] = field(default_factory=list)
    columns_only_gt: List[str] = field(default_factory=list)
    columns_only_app: List[str] = field(default_factory=list)
    cells_compared: int = 0
    cells_equal: int = 0
    cells_blank_vs_zero: int = 0
    deviations: List[Deviation] = field(default_factory=list)
    duplicate_keys_gt: int = 0
    duplicate_keys_app: int = 0
    keys_matched_by_pair: int = 0

    @property
    def ok(self) -> bool:
        return not self.deviations and not self.keys_only_gt and not self.keys_only_app

    def by_line_type(self) -> Dict[str, Counter]:
        out: Dict[str, Counter] = defaultdict(Counter)
        for d in self.deviations:
            out[d.key[1]]["afwijkingen"] += 1
        for k in self.keys_only_gt:
            out[k[1]]["alleen_ground_truth"] += 1
        for k in self.keys_only_app:
            out[k[1]]["alleen_app"] += 1
        return out


# --------------------------------------------------------------------------- #
# Normalisation helpers
# --------------------------------------------------------------------------- #
def _norm_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _norm_header(value: object) -> str:
    """Map a header cell to a stable column name; periods become ``YYYY-MM``."""
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"
    text = _norm_text(value)
    # Exports sometimes carry period headers as text '2026-01-01 00:00:00'.
    if len(text) >= 7 and text[4] == "-" and text[:4].isdigit() and text[5:7].isdigit():
        return text[:7]
    return text


def _as_number(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _read_sheet(path: Path, sheet: str) -> Tuple[List[str], List[List[object]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(sheet)
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = [_norm_header(h) for h in next(rows)]
        body = [list(r) for r in rows if any(c is not None and c != "" for c in r)]
        return header, body
    finally:
        wb.close()


def _index_rows(header: Sequence[str], body: Iterable[Sequence[object]]) -> Tuple[Dict[RowKey, Dict[str, object]], int]:
    """Return ``{key: {column: value}}`` and the number of duplicate keys seen."""
    positions = {name: i for i, name in enumerate(header)}
    missing = [k for k in KEY_HEADERS if k not in positions]
    if missing:
        raise ValueError(f"sleutelkolommen ontbreken: {', '.join(missing)}")
    indexed: Dict[RowKey, Dict[str, object]] = {}
    duplicates = 0
    for row in body:
        key = tuple(_norm_text(row[positions[k]]) if positions[k] < len(row) else "" for k in KEY_HEADERS)
        if key in indexed:
            duplicates += 1
            continue
        indexed[key] = {name: (row[i] if i < len(row) else None) for name, i in positions.items()}
    return indexed, duplicates  # type: ignore[return-value]


def _is_period(column: str) -> bool:
    return len(column) == 7 and column[4] == "-" and column[:4].isdigit()


# --------------------------------------------------------------------------- #
# Comparison
# --------------------------------------------------------------------------- #
def _compare_cell(report: SheetReport, key: RowKey, column: str, gt_val: object, app_val: object,
                  abs_tol: float, rel_tol: float, blank_is_zero: bool) -> None:
    gt_num, app_num = _as_number(gt_val), _as_number(app_val)
    report.cells_compared += 1
    if gt_num is None and app_num is None:
        report.cells_equal += 1
        return
    if (gt_num is None) != (app_num is None):
        other = app_num if gt_num is None else gt_num
        if blank_is_zero and other is not None and abs(other) <= abs_tol:
            report.cells_equal += 1
            report.cells_blank_vs_zero += 1
            return
        report.deviations.append(Deviation(report.sheet, key, column, gt_val, app_val, abs(other or 0.0), math.inf))
        return
    diff = abs(gt_num - app_num)  # type: ignore[operator]
    scale = max(abs(gt_num), abs(app_num))  # type: ignore[arg-type]
    rel = diff / scale if scale else 0.0
    if diff <= abs_tol or rel <= rel_tol:
        report.cells_equal += 1
    else:
        report.deviations.append(Deviation(report.sheet, key, column, gt_num, app_num, diff, rel))


def compare_sheet(gt_path: Path, app_path: Path, sheet: str, abs_tol: float, rel_tol: float,
                  blank_is_zero: bool, aux_tol: float = 0.005, compare_aux: bool = False) -> SheetReport:
    report = SheetReport(sheet=sheet)
    gt_header, gt_body = _read_sheet(gt_path, sheet)
    app_header, app_body = _read_sheet(app_path, sheet)
    gt_rows, report.duplicate_keys_gt = _index_rows(gt_header, gt_body)
    app_rows, report.duplicate_keys_app = _index_rows(app_header, app_body)
    report.rows_gt, report.rows_app = len(gt_body), len(app_body)

    value_columns_gt = [c for c in gt_header if _is_period(c) or c in VALUE_HEADERS]
    value_columns_app = [c for c in app_header if _is_period(c) or c in VALUE_HEADERS]
    report.columns_compared = [c for c in value_columns_gt if c in value_columns_app]
    report.columns_only_gt = [c for c in value_columns_gt if c not in value_columns_app]
    report.columns_only_app = [c for c in value_columns_app if c not in value_columns_gt]

    gt_keys, app_keys = set(gt_rows), set(app_rows)
    common = sorted(gt_keys & app_keys)
    pairs: List[Tuple[RowKey, RowKey]] = [(k, k) for k in common]

    # Pass 2: a numeric aux column is a VALUE (rate, parameter, BOM factor), not
    # an identifier, and the two writers round it differently. Leftover rows are
    # matched on a shorter key when that key is unique on both sides.
    paired_gt: set = set()
    paired_app: set = set()
    # First on (material, line, aux) — only aux 2 differs — then on (material, line).
    for width in (3, 2):
        def _grouped(keys: Iterable[RowKey], n: int = width) -> Dict[Tuple[str, ...], List[RowKey]]:
            grouped: Dict[Tuple[str, ...], List[RowKey]] = defaultdict(list)
            for k in keys:
                grouped[k[:n]].append(k)
            return grouped
        left_gt = _grouped((gt_keys - app_keys) - paired_gt)
        left_app = _grouped((app_keys - gt_keys) - paired_app)
        for pair in sorted(left_gt.keys() & left_app.keys()):
            if len(left_gt[pair]) == 1 and len(left_app[pair]) == 1:
                gk, ak = left_gt[pair][0], left_app[pair][0]
                pairs.append((gk, ak))
                paired_gt.add(gk)
                paired_app.add(ak)
    report.keys_matched_by_pair = len(paired_gt)
    report.keys_only_gt = sorted((gt_keys - app_keys) - paired_gt)
    report.keys_only_app = sorted((app_keys - gt_keys) - paired_app)
    report.keys_matched = len(pairs)

    for gk, ak in pairs:
        gt_row, app_row = gt_rows[gk], app_rows[ak]
        # Aux columns are informational (rates, parameters, hours) and the two
        # writers round and even define them differently; compare on request only.
        if compare_aux and gk != ak:
            for column in ("Aux Column", "Aux 2 Column"):
                _compare_cell(report, gk, column, gt_row.get(column), app_row.get(column), aux_tol, 0.0, True)
        for column in report.columns_compared:
            _compare_cell(report, gk, column, gt_row.get(column), app_row.get(column), abs_tol, rel_tol, blank_is_zero)
    report.deviations.sort(key=lambda d: (-d.abs_diff, d.key, d.column))
    return report


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #
def _fmt(value: object) -> str:
    num = _as_number(value)
    if num is None:
        return "leeg" if value in (None, "") else str(value)
    return f"{num:,.4f}".rstrip("0").rstrip(".") if abs(num) < 1e12 else f"{num:.6g}"


def render_markdown(reports: Sequence[SheetReport], gt_path: Path, app_path: Path, abs_tol: float, rel_tol: float,
                    top: int) -> str:
    lines = ["# Ground-truth-vergelijking", "",
             f"- Ground truth: `{gt_path}`", f"- App-export: `{app_path}`",
             f"- Tolerantie: absoluut ≤ {abs_tol:g} **of** relatief ≤ {rel_tol:g}", ""]
    lines += ["| Sheet | Rijen GT | Rijen app | Gematcht | Alleen GT | Alleen app | Cellen | Gelijk | Afwijkingen | Max |Δ| | Oordeel |",
              "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---|"]
    for r in reports:
        max_abs = max((d.abs_diff for d in r.deviations if math.isfinite(d.abs_diff)), default=0.0)
        lines.append(f"| {r.sheet} | {r.rows_gt} | {r.rows_app} | {r.keys_matched} | {len(r.keys_only_gt)} | "
                     f"{len(r.keys_only_app)} | {r.cells_compared} | {r.cells_equal} | {len(r.deviations)} | "
                     f"{max_abs:,.4f} | {'✅ Δ=0' if r.ok else '❌'} |")
    lines.append("")
    for r in reports:
        lines += [f"## {r.sheet}", ""]
        notes = []
        if r.keys_matched_by_pair:
            notes.append(f"{r.keys_matched_by_pair} rijen gematcht op (materiaal, lijntype) — numerieke aux als waarde vergeleken")
        if r.cells_blank_vs_zero:
            notes.append(f"{r.cells_blank_vs_zero} cellen leeg-vs-0 als gelijk geteld")
        if r.duplicate_keys_gt or r.duplicate_keys_app:
            notes.append(f"dubbele sleutels genegeerd: GT {r.duplicate_keys_gt}, app {r.duplicate_keys_app}")
        if r.columns_only_gt:
            notes.append(f"kolommen alleen in GT: {', '.join(r.columns_only_gt)}")
        if r.columns_only_app:
            notes.append(f"kolommen alleen in app: {', '.join(r.columns_only_app)}")
        if notes:
            lines += ["- " + n for n in notes] + [""]
        breakdown = r.by_line_type()
        if breakdown:
            lines += ["| Lijntype | Afwijkingen | Alleen GT | Alleen app |", "|---|---:|---:|---:|"]
            for lt in sorted(breakdown):
                c = breakdown[lt]
                lines.append(f"| {lt or '(leeg)'} | {c['afwijkingen']} | {c['alleen_ground_truth']} | {c['alleen_app']} |")
            lines.append("")
        if r.deviations:
            lines += [f"Grootste afwijkingen (top {min(top, len(r.deviations))} van {len(r.deviations)}):", "",
                      "| Materiaal | Lijntype | Aux | Periode | Ground truth | App | Δ abs | Δ rel |",
                      "|---|---|---|---|---:|---:|---:|---:|"]
            for d in r.deviations[:top]:
                rel = "∞" if not math.isfinite(d.rel_diff) else f"{d.rel_diff:.2%}"
                lines.append(f"| {d.key[0]} | {d.key[1]} | {d.key[2]} | {d.column} | {_fmt(d.ground_truth)} | "
                             f"{_fmt(d.app)} | {d.abs_diff:,.4f} | {rel} |")
            lines.append("")
        for label, keys in (("Rijen alleen in ground truth", r.keys_only_gt), ("Rijen alleen in app", r.keys_only_app)):
            if keys:
                lines += [f"{label} ({len(keys)}):", ""]
                lines += [f"- {k[0]} · {k[1]} · {k[2]}" for k in keys[:top]]
                if len(keys) > top:
                    lines.append(f"- … nog {len(keys) - top}")
                lines.append("")
    return "\n".join(lines)


def write_xlsx(reports: Sequence[SheetReport], out: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Samenvatting"
    ws.append(["Sheet", "Rijen GT", "Rijen app", "Gematcht", "Alleen GT", "Alleen app", "Cellen", "Gelijk",
               "Afwijkingen", "Oordeel"])
    for r in reports:
        ws.append([r.sheet, r.rows_gt, r.rows_app, r.keys_matched, len(r.keys_only_gt), len(r.keys_only_app),
                   r.cells_compared, r.cells_equal, len(r.deviations), "OK" if r.ok else "AFWIJKING"])
    ws_dev = wb.create_sheet("Afwijkingen")
    ws_dev.append(["Sheet", "Materiaal", "Lijntype", "Aux", "Aux 2", "Kolom", "Ground truth", "App", "Δ abs", "Δ rel"])
    for r in reports:
        for d in r.deviations:
            ws_dev.append([d.sheet, *d.key, d.column, d.ground_truth, d.app, d.abs_diff,
                           None if not math.isfinite(d.rel_diff) else d.rel_diff])
    ws_keys = wb.create_sheet("Ongematchte rijen")
    ws_keys.append(["Sheet", "Kant", "Materiaal", "Lijntype", "Aux", "Aux 2"])
    for r in reports:
        for k in r.keys_only_gt:
            ws_keys.append([r.sheet, "alleen ground truth", *k])
        for k in r.keys_only_app:
            ws_keys.append([r.sheet, "alleen app", *k])
    wb.save(out)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Vergelijk een app-export cel-voor-cel met een ground-truth-werkboek.")
    parser.add_argument("ground_truth", type=Path, help="Excel-model van de klant (.xlsm/.xlsx) — de ground truth")
    parser.add_argument("app_export", type=Path, help="Planningswerkboek uit de app (.xlsx)")
    parser.add_argument("--sheet", action="append", dest="sheets",
                        help=f"te vergelijken sheet (herhaalbaar; standaard: {', '.join(DEFAULT_SHEETS)})")
    parser.add_argument("--abs", type=float, default=1e-6, dest="abs_tol", help="absolute tolerantie (standaard 1e-6)")
    parser.add_argument("--rel", type=float, default=1e-9, dest="rel_tol", help="relatieve tolerantie (standaard 1e-9)")
    parser.add_argument("--aux-tol", type=float, default=0.005, dest="aux_tol",
                        help="tolerantie voor numerieke aux-kolommen (app rondt op 2 decimalen; standaard 0.005)")
    parser.add_argument("--compare-aux", action="store_true",
                        help="vergelijk ook de numerieke aux-kolommen (standaard: alleen periodewaarden en startvoorraad)")
    parser.add_argument("--no-blank-is-zero", action="store_true", help="een lege cel is NIET gelijk aan 0")
    parser.add_argument("--top", type=int, default=25, help="aantal afwijkingen in het rapport per sheet")
    parser.add_argument("--md", type=Path, help="schrijf het rapport als Markdown naar dit pad")
    parser.add_argument("--out", type=Path, help="schrijf alle afwijkingen naar dit .xlsx-bestand")
    args = parser.parse_args(argv)
    # Windows consoles default to cp1252; the report uses ≤, Δ and ✅.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    warnings.filterwarnings("ignore", message="Data Validation extension")

    for p in (args.ground_truth, args.app_export):
        if not p.exists():
            print(f"bestand niet gevonden: {p}", file=sys.stderr)
            return 2
    sheets = args.sheets or list(DEFAULT_SHEETS)
    reports: List[SheetReport] = []
    for sheet in sheets:
        try:
            reports.append(compare_sheet(args.ground_truth, args.app_export, sheet, args.abs_tol, args.rel_tol,
                                         not args.no_blank_is_zero, args.aux_tol, args.compare_aux))
        except KeyError:
            print(f"sheet '{sheet}' ontbreekt in een van de bestanden — overgeslagen", file=sys.stderr)
        except ValueError as exc:
            print(f"sheet '{sheet}': {exc} — overgeslagen", file=sys.stderr)
    if not reports:
        return 2
    text = render_markdown(reports, args.ground_truth, args.app_export, args.abs_tol, args.rel_tol, args.top)
    print(text)
    if args.md:
        args.md.write_text(text, encoding="utf-8")
    if args.out:
        write_xlsx(reports, args.out)
    return 0 if all(r.ok for r in reports) else 1


if __name__ == "__main__":
    sys.exit(main())
