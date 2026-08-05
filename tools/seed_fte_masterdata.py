#!/usr/bin/env python
"""Vul de app-masterdata (F2-CF) vanuit het Maastricht OEE/FTE-klantmodel.

Eenmalige seed uit `OEE model MTO APEX voorbeeld.xlsx` (site NLX1), zoals
afgesproken in docs/plan-f2-cf-fte-werkbank.md §5.3. Daarna is het Excel-bestand
niet meer nodig: de waarden zijn masterdata en worden in de app of het
masterwerkboek beheerd.

Uitgangspunten
--------------
1. **Het script LEEST het werkboek; het bevat geen overgetypte getallen.**
   Alleen de vertaaltabel tussen de klantnamen (PM01, KM03, MRL, …) en de
   SAP-machinecodes (PML01, PML06, PSS13, …) staat hier hard — die staat
   nergens in het werkboek en is met de OEE-waarden geverifieerd.
2. **Niets wordt dubbel geteld.** Wat de app al berekent (controlekamer,
   truckbelading via ZZZZ_TRUCK01/02, de MAX over crusher+zeef binnen
   ZZ_GROUP01) wordt NIET geseed. Wat de klant wél heeft maar de app niet
   (onderhoud, truckLOSSING, feed door derden) wél.
3. **Wat we niet kunnen weten, zetten we niet aan.** Regels waarvan de
   volumebron in de app nog niet vaststaat, komen als `is_active=False` in de
   database: zichtbaar, met de klantnorm erbij, en met één vinkje te activeren
   zodra de bron bekend is. Verzinnen is geen optie bij klantcijfers.
4. **Droogdraaien is de standaard.** Zonder `--apply` verandert er niets en
   krijg je alleen het rapport. Met `--apply` wordt de store eerst geback-upt.

Gebruik
-------
    python tools/seed_fte_masterdata.py --report %LOCALAPPDATA%\\SOPPlanningEngine\\seed.md
    python tools/seed_fte_masterdata.py --apply
    python tools/seed_fte_masterdata.py --apply --machines-workbook pad/naar/MS_RECONC.xlsm

Een tweede --apply voegt alleen ONTBREKENDE records toe. Bestaande records —
inclusief wat de klant inmiddels heeft aangepast — blijven staan tenzij je
--overwrite-existing meegeeft.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from modules.models import (  # noqa: E402
    BenchmarkThroughput,
    IndirectActivity,
    StaffingNorm,
    derive_effective_fte_hours,
)

DEFAULT_WORKBOOK = Path.home() / 'Documents' / 'OEE model MTO APEX voorbeeld.xlsx'

# Deze seed bevat Maastricht-normen. De zusterrepo's draaien op een eigen
# SOP_APP_DATA_DIR; zonder controle landen ze zo in de Winterswijk-store.
EXPECTED_SITE = 'NLX1'

SHEET_FTE = 'FTE'
SHEET_MODEL = 'OEE Model MST '
SHEET_MES = 'MES_OEE Mills'
SHEET_PEER = 'PEER_Capacity'
SHEET_NORMEN = 'Normen '

# ── Vertaaltabel klantmodel → SAP ───────────────────────────────────────────
# Het klantmodel gebruikt werkvloernamen, de app SAP-machinecodes. Deze tabel
# is geverifieerd op de OEE-waarde: elke regel hieronder heeft in het model
# dezelfde OEE als de SAP-machine in de masterdata (zie --verify-mapping).
MODEL_TO_SAP = {
    'MRL':        ('PSS13', 'Marl Crusher'),
    'MGN':        ('PWP01', 'Mogensen Sieve'),
    'PM01':       ('PML01', 'Pendel Mill 1'),
    'PM02':       ('PML02', 'Pendel Mill 2'),
    'PM03':       ('PML03', 'Pendel Mill 3'),
    'PE24':       ('PML24', 'Roller Mill PE24'),
    'PE15':       ('PML15', 'Production unit 15'),
    'PE16':       ('PML16', 'Production Unit 16'),
    'PE21':       ('PML21', 'Production unit 21'),
    'PE12':       ('PML12', 'Production Unit 12'),
    'PE12C':      ('PML11', 'Production Unit 12 — coating'),
    'KM02':       ('PML05', 'Ball Mill 2 (Kogelmolen 2)'),
    'KM03':       ('PML06', 'Ball Mill 3 (Kogelmolen 3)'),
    'PE08':       ('PPM08', 'Mixing Installation Borga'),
    'PE14':       ('PPM14', 'Filler Preparation'),
    'Pneu':       ('PPM09', 'Pneumex'),
    'TBMA A + C': ('PBA19', 'Bigbag Machine TBMA'),
    'BB PE24':    ('PBA24', 'BB Machine TBMA B'),
    'BB H&B':     ('PBA11', 'Bigbag Machine H&B'),
    'BB PE20':    ('PBA20', 'Bagging Machine PE20'),
    'BB PE25':    ('PBA25', 'Bagging Machine PE25'),
    'ZVM07':      ('PBA07', 'Bagging Machine 7'),
    'PE20':       ('PBA20', 'Bagging Machine PE20'),
    'PE25':       ('PBA25', 'Bagging Machine PE25'),
}

# Werkcenternamen op het MES-blad → SAP-machinecode. Namen zonder eenduidige
# machine (verzamelnamen als 'Malen', 'MTO') staan er bewust niet in; die
# worden overgeslagen en in het rapport genoemd.
MES_WORKCENTER_TO_SAP = {
    'Bagging Machine PE25': 'PBA25',
    'Big Bag H&B': 'PBA11',
    'Big Bag TBMA A': 'PBA19',
    'Big Bag TBMA B': 'PBA24',
    'Big Bag TBMA C': 'PBA19',
    'Coating': 'PML11',
    'Kogelmolen 2': 'PML05',
    'Kogelmolen 3': 'PML06',
    'Mergelbreker': 'PSS13',
    'Mogensen zeef': 'PWP01',
    'PE12': 'PML12',
    'PE15': 'PML15',
    'PE16': 'PML16',
    'PE20': 'PBA20',
    'PE24': 'PML24',
    'Pendelmolen 1': 'PML01',
    'Pendelmolen 2': 'PML02',
    'Pendelmolen 3': 'PML03',
    'Vulstof Voorbereiding': 'PPM14',
    'Zakkenvulmachine7': 'PBA07',
}

# Machinegroepen in de app en hun betekenis in het klantmodel. De coëfficiënt
# '# FTE Staffing' staat in het model op rij 165 en is voor élk blok 1 —
# één operator per draaiend uur. We seeden hem expliciet zodat in de werkbank
# zichtbaar is dat het een KLANTNORM is en niet de default.
GROUP_MEANING = {
    'ZZ_GROUP01': 'Mergelbreker + Mogensen-zeef (model: Crusher / Mogensen Sieve)',
    'ZZ_GROUP02': 'Pendelmolens 1-3 + rollenmolen PE24 (model: PM01/PM02/PM03/PE24)',
    'ZZ_GROUP03': 'Productie-units 15, 16, 21 + breker 21 (model: PE15/PE16/PE21)',
    'ZZ_GROUP04': 'Kogelmolens 2 en 3 + unit 12 (model: KM02/KM03/PE12)',
    'ZZ_GROUP05': 'Mengsinstallatie Borga (model: PE08)',
    'ZZ_GROUP06': 'Vulstofvoorbereiding (model: PE14)',
    'ZZ_GROUP07': 'Unit 12 coating (model: PE12C)',
    'ZZ_GROUP08': 'Pneumex (model: Pneu)',
    'ZZZ_PACKGROUP01': 'Big-bag- en zaklijnen (model: BigBag-blok)',
    'ZZZ_PACKGROUP02': 'Zakkenvulmachine 7 (model: Small Bags-blok)',
}

# Korte, leesbare naam per groep. De groepsmaterialen in de masterdata hebben
# geen naam (de cel is leeg en komt als de tekst 'nan' binnen), waardoor de
# planningtabel, de exports én de werkbank rijen tonen die "nan" heten. Deze
# namen zetten we op het groepsmateriaal zelf, zodat het overal doorwerkt.
GROUP_NAME = {
    'ZZ_GROUP01': 'Breken & zeven',
    'ZZ_GROUP02': 'Pendelmolens',
    'ZZ_GROUP03': 'Productie-units 15/16/21',
    'ZZ_GROUP04': 'Kogelmolens',
    'ZZ_GROUP05': 'Mengen (Borga)',
    'ZZ_GROUP06': 'Vulstofvoorbereiding',
    'ZZ_GROUP07': 'Coating',
    'ZZ_GROUP08': 'Pneumex',
    'ZZ_GROUP09': 'Molengroep 9 (ongebruikt)',
    'ZZZ_PACKGROUP01': 'Big-bag- en zaklijnen',
    'ZZZ_PACKGROUP02': 'Zakkenvulmachine 7',
    'ZZZ_PACKGROUP03': 'Verpakgroep 3 (ongebruikt)',
    'ZZZ_PACKGROUP04': 'Verpakgroep 4 (ongebruikt)',
}


def _num(value):
    """Celwaarde → float, of None als de cel leeg of tekst is ('x' = n.v.t.)."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


def _text(value) -> str:
    """Celtekst, met regeleinden weggevouwen — een harde return in een cel
    ('TBMA\\nH&B') breekt anders elke markdown-tabelregel in het rapport."""
    if value is None:
        return ''
    return ' '.join(str(value).split())


# ── blad 'FTE' ──────────────────────────────────────────────────────────────


def _rows_by_label(ws, column: int = 2) -> dict:
    """{label in kleine letters: rijnummer} voor een kolom met labels.

    Vaste celadressen braken zodra de klant één rij invoegde: F13 wees dan
    naar een lege cel (uren per FTE = 0, door hydratie stil 1492) en de
    ploegurenrijen schoven op, waardoor élke 3-ploegenmachine 347 in plaats
    van 520 beschikbare uren kreeg. Het label verschuift niet.
    """
    labels = {}
    duplicates = {}
    for row in range(1, ws.max_row + 1):
        text = _text(ws.cell(row=row, column=column).value).lower()
        if not text:
            continue
        if text in labels:
            duplicates.setdefault(text, [labels[text]]).append(row)
        else:
            labels[text] = row
    # Een dubbel label is dodelijk: we zouden stil de EERSTE lezen. Voor de
    # bruto->netto-velden vangen de afleidingscontroles dat nog af, maar voor
    # de bezettingsgraad en de ploeguren is er niets dat het merkt.
    labels['__duplicates__'] = duplicates
    return labels


class SeedError(RuntimeError):
    """Het werkboek levert niet wat de seed nodig heeft. Luid falen: stil
    doorschrijven zet onzin in de klantdatabase."""


def read_fte(ws) -> tuple[dict, float, dict, list[str]]:
    """FTE-parameters, het eindgetal, de ploeguren en de controleregels."""
    labels = _rows_by_label(ws)

    duplicates = labels.get('__duplicates__') or {}

    def _row(label: str) -> int:
        key = label.lower()
        if key in duplicates:
            rows = ', '.join(str(r) for r in duplicates[key])
            raise SeedError(
                f'Blad "FTE": "{label}" staat op meerdere rijen ({rows}). '
                f'De seed leest op label, dus welke bedoeld is valt niet af '
                f'te leiden. Maak de labels uniek.')
        row = labels.get(key)
        if row is None:
            raise SeedError(
                f'Blad "FTE": regel "{label}" niet gevonden. Is het werkboek '
                f'van vorm veranderd? De seed leest op label, niet op celadres.')
        return row

    def _cell(label: str, column: int):
        return _num(ws.cell(row=_row(label), column=column).value)

    gross = _cell('Effective days', 6) or 0.0
    leave = abs(_cell('Holiday', 6) or 0.0)
    adv = abs(_cell('ATV', 6) or 0.0)
    holiday = abs(_cell('Public holidays in workweek', 6) or 0.0)
    after_holidays = _cell('Working hours after holidays', 6) or 0.0
    illness = _cell('Sick leave', 3) or 0.0
    training = _cell('Training', 3) or 0.0
    effective = _cell('Working hours per FTE per yr', 6) or 0.0
    occupancy = _cell('Bezettingsgraad', 4) or 0.0

    params = {
        'utilization_rate': occupancy,
        'gross_hours_per_year': gross,
        'leave_hours_per_year': leave,
        'adv_hours_per_year': adv,
        'holiday_hours_per_year': holiday,
        'illness_pct': illness,
        'training_pct': training,
    }

    shift_hours = {}
    for label, name in (('2-ploegen', '2-shift system'),
                        ('3-ploegen', '3-shift system'),
                        ('24/7', '24/7 production')):
        annual = _cell(label, 3)
        if not annual or annual <= 0:
            raise SeedError(f'Blad "FTE": jaaruren bij "{label}" ontbreken of zijn 0.')
        # Plausibiliteitsbereik: een ploegensysteem ligt tussen ~1.200 (één
        # ploeg) en 8.784 (24/7 in een schrikkeljaar) uur per jaar. Zonder
        # deze grens leest een verschoven of verkeerd label stil een getal dat
        # de beschikbare uren van elke machine in dat systeem scheeftrekt.
        if not 1000 <= annual <= 8784:
            raise SeedError(
                f'Blad "FTE": {annual:g} jaaruren bij "{label}" is onmogelijk '
                f'(verwacht tussen 1.000 en 8.784). Lees het blad na.')
        shift_hours[name] = annual / 12

    # Blokkerende plausibiliteitscontroles. Een 0 hier komt via hydratie stil
    # als 1492 terug en verschuift elk FTE-getal.
    if gross <= 0:
        raise SeedError('Blad "FTE": bruto uren per jaar is 0 of ontbreekt.')
    if effective <= 0:
        raise SeedError('Blad "FTE": uren per FTE per jaar is 0 of ontbreekt.')
    if not 0 < occupancy <= 1:
        raise SeedError(f'Blad "FTE": bezettingsgraad {occupancy} valt buiten (0, 1].')

    checks = []
    base = gross - leave - adv - holiday
    if abs(base - after_holidays) >= 0.01:
        raise SeedError(
            f'Blad "FTE": bruto {gross:g} - verlof {leave:g} - ADV {adv:g} - '
            f'feestdagen {holiday:g} = {base:g}, maar het werkboek zegt '
            f'{after_holidays:g}. Lees het blad na voordat je seedt.')
    checks.append(f'bruto {gross:g} - verlof {leave:g} - ADV {adv:g} - feestdagen '
                  f'{holiday:g} = {base:g} (werkboek = {after_holidays:g}) OK')
    derived = derive_effective_fte_hours(params)
    if abs(derived - effective) >= 0.01:
        raise SeedError(
            f'Blad "FTE": de afleiding geeft {derived:.2f} maar het werkboek voert '
            f'{effective:.2f} als uren per FTE. Controleer ziekte-/trainingspercentage.')
    checks.append(f'{base:g} - {illness:.0%} - {training:.0%} van diezelfde basis '
                  f'= {derived:.2f} (werkboek = {effective:.2f}) OK')
    compounded = base * (1 - illness) * (1 - training)
    checks.append(f'ter vergelijking: stapelend gerekend zou {compounded:.2f} geven - '
                  f'{compounded - effective:+.2f} uur per FTE per jaar')
    return params, effective, shift_hours, checks


# ── blad 'OEE Model MST ' ───────────────────────────────────────────────────


def read_model_staffing(ws) -> tuple[float, list[str]]:
    """De '# FTE Staffing'-coëfficiënt uit rij 165; alle blokken moeten gelijk zijn."""
    values = [v for v in (_num(c.value) for c in ws[165]) if v is not None]
    notes = [f'blokken met een bemensingsnorm: {len(values)}; waarden: '
             f'{sorted(set(values))}']
    coefficient = values[0] if values else 1.0
    if len(set(values)) > 1:
        notes.append('LET OP: niet elk blok heeft dezelfde norm — hier is de eerste '
                     'gebruikt; controleer het werkboek.')
    return coefficient, notes


def read_model_machines(ws) -> dict:
    """{klantnaam: {'oee': {...}, 'capacity': {...}}} uit de kolomkoppen.

    Rij 10 draagt de naam, rij 7 de OEE en rij 8 de capaciteit in t/uur. De
    Processing-blokken herhalen dezelfde machines, dus we verzamelen per naam
    de set waarden en kunnen zien of de blokken het eens zijn.
    """
    names = {}
    for column in range(2, ws.max_column + 1):
        name = _text(ws.cell(row=10, column=column).value)
        if not name or name in ('Product', 'Volume', 'Recept', 'Processing',
                                'Capacity', 'OEE'):
            continue
        entry = names.setdefault(name, {'oee': set(), 'capacity': set()})
        oee = _num(ws.cell(row=7, column=column).value)
        capacity = _num(ws.cell(row=8, column=column).value)
        if oee is not None:
            entry['oee'].add(round(oee, 4))
        if capacity is not None:
            entry['capacity'].add(round(capacity, 4))
    return names


def verify_mapping(model_machines: dict, sap_machines: dict) -> list[str]:
    """Vergelijk de OEE van elke klantmachine met die van de gekoppelde
    SAP-machine. Gelijke OEE's zijn het bewijs dat de vertaaltabel klopt."""
    lines = ['| klantmodel | SAP-code | OEE model | OEE masterdata | t/u model | oordeel |',
             '|---|---|---:|---:|---:|---|']
    for name in sorted(model_machines):
        entry = model_machines[name]
        target = MODEL_TO_SAP.get(name)
        oee_model = sorted(entry['oee'])
        capacity = sorted(entry['capacity'])
        oee_text = '/'.join(f'{v:g}' for v in oee_model) or '—'
        cap_text = '/'.join(f'{v:g}' for v in capacity) or '—'
        if target is None:
            lines.append(f'| {name} | — | {oee_text} | — | {cap_text} | GEEN KOPPELING |')
            continue
        code = target[0]
        machine = sap_machines.get(code)
        if machine is None:
            lines.append(f'| {name} | {code} | {oee_text} | ontbreekt | {cap_text} '
                         f'| machine niet in de masterdata |')
            continue
        oee_sap = round(float(machine.get('oee') or 0.0), 4)
        match = any(abs(v - oee_sap) < 0.005 for v in oee_model) if oee_model else False
        lines.append(f'| {name} | {code} | {oee_text} | {oee_sap:g} | {cap_text} '
                     f'| {"OK" if match else "OEE WIJKT AF"} |')
    return lines


def read_truck_rows(ws) -> tuple[list[dict], list[dict]]:
    """Truckbelading (185-188) en -lossing (192-193)."""
    def _rows(start, end):
        out = []
        for row in range(start, end + 1):
            label = _text(ws[f'D{row}'].value)
            if not label or label.lower() == 'total':
                continue
            out.append({
                'row': row,
                'label': label,
                'share': _num(ws[f'C{row}'].value),
                'contractor': _text(ws[f'C{row}'].value) if _num(ws[f'C{row}'].value) is None else '',
                'hours_per_truck': _num(ws[f'E{row}'].value),
                'volume': _num(ws[f'F{row}'].value),
                'tons_per_truck': _num(ws[f'G{row}'].value),
                'trucks': _num(ws[f'H{row}'].value),
                'fte': _num(ws[f'I{row}'].value),
            })
        return out
    return _rows(185, 188), _rows(192, 193)


def read_feed_rows(ws) -> list[dict]:
    """Marl (stort) en Mill Feed — handling door derden, rijen 179-180."""
    out = []
    for row in (179, 180):
        label = _text(ws[f'D{row}'].value)
        if not label:
            continue
        out.append({
            'row': row,
            'label': label,
            'contractor': _text(ws[f'C{row}'].value),
            'tons_per_hour': _num(ws[f'E{row}'].value),
            'volume': _num(ws[f'F{row}'].value),
            'hours': _num(ws[f'G{row}'].value),
            'fte': _num(ws[f'I{row}'].value),
            'remark': _text(ws[f'J{row}'].value),
        })
    return out


def read_maintenance(ws) -> dict:
    return {
        'label': _text(ws['B198'].value),
        'fte': _num(ws['D198'].value),
        'unit': _text(ws['E198'].value),
        'machine_hours_total': _num(ws['D200'].value),
        'machine_hours_per_fte': _num(ws['E200'].value),
        'opex_total': _num(ws['D201'].value),
        'opex_per_fte': _num(ws['E201'].value),
    }


def read_control_room(ws) -> dict:
    return {'fte': _num(ws['I175'].value), 'total_fte': _num(ws['I195'].value)}


# ── benchmarkbladen ─────────────────────────────────────────────────────────


def read_mes(ws) -> tuple[dict, list[str]]:
    """Gemeten OEE per werkcenter → {sap_code: (oee, doel, opmerking)}."""
    measured, skipped = {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _text(row[0])
        if not name or name.lower() == 'total' or name.startswith('Applied filters'):
            continue
        code = MES_WORKCENTER_TO_SAP.get(name)
        if code is None:
            skipped.append(name)
            continue
        oee = _num(row[1]) or 0.0
        target = _num(row[3]) if len(row) > 3 else None
        remark = _text(row[4]) if len(row) > 4 else ''
        if oee <= 0:
            skipped.append(f'{name} (geen meting: OEE 0)')
            continue
        measured[code] = (oee, target, remark)
    return measured, skipped


def _between_parentheses(text: str) -> str:
    if '(' in text and ')' in text:
        return text[text.rindex('(') + 1:text.rindex(')')].strip()
    return ''


def read_peer(ws) -> tuple[list[dict], list[str]]:
    """PEER-doorzet per machine en per (machine, materiaalnummer).

    Het blad is een draaitabel: de installatienaam staat alleen op de eerste
    regel van elk blok en is daaronder samengevoegd/leeg. We vullen hem naar
    beneden door — zonder dat leest élke productregel als 'geen machine' en
    valt de hele tabel weg.
    """
    rows, skipped = [], []
    machine_text = ''
    for values in ws.iter_rows(min_row=2, values_only=True):
        cell = _text(values[1])
        if cell:
            machine_text = cell
        product_text = _text(values[2])
        throughput = _num(values[3])
        if not machine_text or machine_text.lower() == 'total' or throughput is None:
            continue
        code = _between_parentheses(machine_text)
        if not code:
            skipped.append(f'{machine_text} (geen machinecode tussen haakjes)')
            continue
        if product_text.lower() == 'total':
            rows.append({'machine_code': code, 'material_number': '',
                         'throughput': throughput, 'machine_label': machine_text,
                         'product_label': 'installatietotaal'})
            continue
        material = _between_parentheses(product_text)
        if not material:
            skipped.append(f'{machine_text} / {product_text} (geen materiaalnummer)')
            continue
        rows.append({'machine_code': code, 'material_number': material,
                     'throughput': throughput,
                     'machine_label': machine_text, 'product_label': product_text})
    return rows, skipped


def read_normen(ws) -> list[dict]:
    """Normen-tabel: productie-unit × product met MES-, PEER- en OEE-normen.

    Blijft BUITEN de database: de productkolom bevat handelsnamen (Ank125,
    SG63), geen materiaalnummers, dus een koppeling aan de app-materialen kan
    niet zonder de klant. Wel opgenomen in het rapport, zodat de vraag concreet
    gesteld kan worden.
    """
    out, unit, avail = [], '', None
    for row in ws.iter_rows(min_row=5, values_only=True):
        unit_cell = _text(row[0])
        if unit_cell:
            unit, avail = unit_cell, _num(row[6])
        product = _text(row[2])
        if not product or not unit:
            continue
        out.append({
            'unit': unit, 'product': product,
            'mes': _num(row[3]), 'peer': _num(row[4]),
            'performance': _num(row[5]),
            'availability': _num(row[6]) if _num(row[6]) is not None else avail,
            'oee': _num(row[7]),
        })
    return out


# ── seed opbouwen ───────────────────────────────────────────────────────────


def _hours_per_period(annual_hours: float) -> float:
    return annual_hours / 12


def build_seed(workbook_path: Path, sap_machines: dict | None = None) -> tuple[dict, list[str], list[str]]:
    """Bouw de te seeden datasets + rapportregels + openstaande vragen."""
    import openpyxl

    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    report: list[str] = []
    questions: list[str] = []

    # 1. FTE-parameters ------------------------------------------------------
    params, effective_hours, shift_hours, checks = read_fte(wb[SHEET_FTE])
    report.append('## 1. FTE-parameters (blad `FTE`)\n')
    report.append('| parameter | waarde | betekenis |')
    report.append('|---|---:|---|')
    # Labels, geen celadressen: de seed leest op label (een ingevoegde rij
    # verschuift de cellen wel maar het label niet), dus adressen noemen zou
    # de lezer naar de verkeerde cel sturen.
    report.append(f'| bruto uren per jaar | {params["gross_hours_per_year"]:g} | regel "Effective days" |')
    report.append(f'| verlof | {params["leave_hours_per_year"]:g} | regel "Holiday" |')
    report.append(f'| ADV | {params["adv_hours_per_year"]:g} | regel "ATV" |')
    report.append(f'| feestdagen in de werkweek | {params["holiday_hours_per_year"]:g} | regel "Public holidays in workweek" |')
    report.append(f'| ziekteverzuim | {params["illness_pct"]:.0%} | regel "Sick leave" |')
    report.append(f'| training | {params["training_pct"]:.0%} | regel "Training" |')
    report.append(f'| **effectieve uren per FTE per jaar** | **{effective_hours:.2f}** | regel "Working hours per FTE per yr" — dit getal rekent |')
    report.append(f'| bezettingsgraad | {params["utilization_rate"]:.0%} | regel "Bezettingsgraad" |')
    report.append('')
    report.append('Controle van de rekenregel:')
    for line in checks:
        report.append(f'- {line}')
    report.append('')
    report.append('Ploeguren per maand (jaaruren ÷ 12, gelezen op de regels '
                  '"2-ploegen", "3-ploegen" en "24/7"): '
                  + ', '.join(f'{k} = {v:.1f}' for k, v in shift_hours.items()))
    report.append('')

    # 2. Bemensingsnormen ----------------------------------------------------
    model_ws = wb[SHEET_MODEL]
    coefficient, staffing_notes = read_model_staffing(model_ws)
    staffing_norms = {}
    for group_id, meaning in GROUP_MEANING.items():
        staffing_norms[group_id] = asdict(StaffingNorm(
            code=group_id,
            operators_per_hour=coefficient,
            scope='group',
            function_group='operators',
            description=f'Klantmodel "# FTE Staffing" = {coefficient:g} — {meaning}',
        ))
    report.append('## 2. Bemensingsnormen (blad `OEE Model MST `, rij 165)\n')
    report.append(f'Het klantmodel rekent met **{coefficient:g} operator per draaiend uur** voor élk blok. '
                  'Dat is hetzelfde getal dat Line 12 vandaag als default gebruikt, dus de FTE-cijfers '
                  'veranderen hier niet van — de norm wordt alleen expliciet en zichtbaar in de werkbank.')
    for note in staffing_notes:
        report.append(f'- {note}')
    report.append('')
    report.append('| groep | operators/draaiuur | betekenis in het klantmodel |')
    report.append('|---|---:|---|')
    for group_id, meaning in GROUP_MEANING.items():
        report.append(f'| {group_id} | {coefficient:g} | {meaning} |')
    report.append('')

    # 2b. Controle van de vertaaltabel klantnaam → SAP-code ------------------
    model_machines = read_model_machines(model_ws)
    report.append('### 2b. Klopt de koppeling klantnaam → SAP-machinecode?')
    report.append('')
    if sap_machines:
        report.append('Het klantmodel noemt machines bij hun werkvloernaam, de app bij hun '
                      'SAP-code. De OEE is de controle: die staat in beide bronnen en moet '
                      'gelijk zijn.')
        report.append('')
        report.extend(verify_mapping(model_machines, sap_machines))
        mismatches = [line for line in verify_mapping(model_machines, sap_machines)
                      if 'WIJKT AF' in line or 'GEEN KOPPELING' in line or 'niet in de' in line]
        report.append('')
        if mismatches:
            report.append(f'{len(mismatches)} regel(s) vragen aandacht — zie het oordeel hierboven.')
            questions.append('Machinekoppeling: controleer de regels met "OEE WIJKT AF" of '
                             '"GEEN KOPPELING" in tabel 2b.')
        else:
            report.append('Alle klantmachines koppelen aan een SAP-machine met dezelfde OEE.')
    else:
        report.append('Geen machines in de masterdata om tegen te controleren — draai met '
                      '`--machines-workbook` om de vertaaltabel te laten verifiëren.')
    report.append('')

    # 3. Indirecte activiteiten ---------------------------------------------
    loading, unloading = read_truck_rows(model_ws)
    feed = read_feed_rows(model_ws)
    maintenance = read_maintenance(model_ws)
    control_room = read_control_room(model_ws)
    activities: dict[str, dict] = {}
    model_fte: dict[str, float] = {}

    slug = {'Container': 'CONTAINER', 'Pallet/BB': 'PALLET', 'Bulk (PO)': 'BULK_PO',
            'Bulk (WS)': 'BULK_WS', 'Containers': 'CONTAINER'}

    for item in loading:
        key = f'TRUCK_LOAD_{slug.get(item["label"], item["label"].upper())}'
        model_fte[key] = item['fte'] or 0.0
        activities[key] = asdict(IndirectActivity(
            activity_id=key,
            name=f'Truck laden — {item["label"]}',
            driver='per_truck',
            hours_per_unit=item['hours_per_truck'] or 0.0,
            tons_per_truck=item['tons_per_truck'] or 0.0,
            function_group='operators',
            volume_line='01. Demand forecast',
            is_active=False,
            description=(
                f'Klantmodel rij {item["row"]}: {item["volume"]:,.0f} t ÷ {item["tons_per_truck"]:g} t/truck '
                f'= {item["trucks"]:,.0f} trucks × {item["hours_per_truck"]:g} u = {item["fte"]:.3f} FTE. '
                'UIT: de app rekent truckbelading al via ZZZZ_TRUCK01/ZZZZ_TRUCK02 in Line 12 — '
                'aanzetten zou dubbel tellen. Staat hier als klantnorm en als basis voor een '
                'fijnere opsplitsing van de truckmaterialen.'),
        ))

    for item in unloading:
        key = f'TRUCK_UNLOAD_{slug.get(item["label"], item["label"].upper())}'
        model_fte[key] = item['fte'] or 0.0
        tons = item['tons_per_truck'] or 0.0
        activities[key] = asdict(IndirectActivity(
            activity_id=key,
            name=f'Truck lossen — {item["label"]}',
            driver='per_truck' if tons > 0 else 'fixed',
            hours_per_unit=item['hours_per_truck'] or 0.0,
            tons_per_truck=tons,
            fte_per_period=(item['fte'] or 0.0) if tons <= 0 else 0.0,
            function_group='operators',
            volume_line='06. Purchase receipt',
            volume_source='',
            is_active=False,
            description=(
                f'Klantmodel rij {item["row"]}: {item["trucks"]:,.0f} trucks × '
                f'{item["hours_per_truck"]:g} u = {item["fte"]:.4f} FTE. '
                'De app kent truckLOSSING nog niet, dus dit is nieuwe informatie. UIT tot de klant '
                'zegt WELKE inkomende materialen per container/pallet binnenkomen — nu aanzetten zou '
                'de lostijd over álle inkoopontvangsten smeren.'),
        ))
        questions.append(
            f'Truck lossen ({item["label"]}): welke inkoopmaterialen komen zo binnen? '
            f'Zonder die afbakening kan de regel niet aan.')

    for item in feed:
        key = 'FEED_MARL' if 'marl' in item['label'].lower() else 'FEED_MILL'
        model_fte[key] = item['fte'] or 0.0
        per_ton = 1 / item['tons_per_hour'] if item['tons_per_hour'] else 0.0
        activities[key] = asdict(IndirectActivity(
            activity_id=key,
            name=f'{item["label"]}' + (f' ({item["contractor"]})' if item['contractor'] else ''),
            driver='per_ton',
            hours_per_unit=per_ton,
            function_group='operators',
            volume_line='06. Production plan',
            volume_source='',
            is_active=False,
            description=(
                f'Klantmodel rij {item["row"]}: {item["volume"]:,.0f} t ÷ {item["tons_per_hour"]:g} t/u '
                f'= {item["hours"]:,.0f} u = {item["fte"]:.3f} FTE'
                + (f'. {item["remark"]}' if item['remark'] else '')
                + (f'. Uitvoering: {item["contractor"]}' if item['contractor'] else '')
                + '. UIT tot de volumebron vaststaat: het klantmodel gebruikt een deelvolume '
                  f'({item["volume"]:,.0f} t van de {control_room["total_fte"] and ""}totale sitedoorzet), '
                  'niet het sitetotaal.'),
        ))
        questions.append(
            f'{item["label"]}: welk deel van het volume drijft deze regel '
            f'({item["volume"]:,.0f} t in het model)? Materiaalnummer of productfamilie graag.')

    model_fte['MAINTENANCE_DIRECT'] = maintenance['fte'] or 0.0
    activities['MAINTENANCE_DIRECT'] = asdict(IndirectActivity(
        activity_id='MAINTENANCE_DIRECT',
        name='Direct onderhoud',
        driver='fixed',
        fte_per_period=maintenance['fte'] or 0.0,
        function_group='maintenance',
        is_active=True,
        description=(
            f'Klantmodel rij 198: "{maintenance["label"]}" = {maintenance["fte"]:g} FTE (vaste bezetting). '
            f'Kengetallen uit rij 200/201: {maintenance["machine_hours_per_fte"]:,.0f} machine-uren per '
            f'onderhouds-FTE en € {maintenance["opex_per_fte"]:,.0f} OPEX per onderhouds-FTE. '
            'AAN: onderhoud zit niet in Line 12, dus dit is echt extra en geen dubbeltelling. '
            'Let op: de "9" is een AANTAL FTE, geen "9 machines per FTE" — het totaal van 32,06 FTE '
            'in rij 195 is exclusief onderhoud.'),
    ))

    report.append('## 3. Indirecte activiteiten\n')
    report.append('| activiteit | driver | aan? | FTE volgens de klant | waarom |')
    report.append('|---|---|:--:|---:|---|')
    for key, item in activities.items():
        first_sentence = item['description'].split('. ')[0]
        aan = 'JA' if item['is_active'] else 'nee'
        report.append(f'| {item["name"]} | {item["driver"]} | {aan} | '
                      f'{model_fte.get(key, 0.0):.3f} | {first_sentence}. |')
    total_seeded = sum(model_fte.values())
    active_seeded = sum(v for k, v in model_fte.items() if activities[k]['is_active'])
    report.append('')
    report.append(f'Samen {total_seeded:.2f} FTE aan klantnormen in de database, waarvan '
                  f'{active_seeded:.2f} FTE nu meetelt in de werkbank. De rest staat uit '
                  '(dubbeltelling of onbekende volumebron) en is met één vinkje aan te zetten.')
    report.append('')
    report.append(f'**Niet geseed — de app doet dit al:** de controlekamer. Het klantmodel komt op '
                  f'{control_room["fte"]:.3f} FTE (rij 175); de app rekent via ZZZZZ_CONTROLROOM exact '
                  f'hetzelfde getal. Een activiteit toevoegen zou het verdubbelen.')
    report.append('')
    report.append(f'**Niet geseed — de app doet dit al:** "Crusher + sieve" (rij 181, {model_ws["I181"].value:.3f} FTE) '
                  'is het MAXIMUM van de breker en de zeef. Die twee zitten in de app samen in ZZ_GROUP01, '
                  'een molengroep, en molengroepen aggregeren al met MAX. Een machinecombinatie aanmaken '
                  'zou dezelfde bezetting nog eens tellen.')
    report.append('')

    # 4. Benchmarks ----------------------------------------------------------
    mes, mes_skipped = read_mes(wb[SHEET_MES])
    peer_rows, _peer_skipped = read_peer(wb[SHEET_PEER])
    benchmarks: dict[str, dict] = {}
    for row in peer_rows:
        key = f'{row["machine_code"]}|{row["material_number"]}'
        oee, target, remark = mes.get(row['machine_code'], (0.0, None, ''))
        benchmarks[key] = asdict(BenchmarkThroughput(
            machine_code=row['machine_code'],
            material_number=row['material_number'],
            peer_t_per_hour=row['throughput'],
            mes_oee=oee,
            note=f'PEER: {row["machine_label"]} / {row["product_label"]}'
                 + (f' · MES-OEE gemeten {oee:.1%}' if oee else '')
                 + (f', doel {target:.0%}' if target else '')
                 + (f' — {remark}' if remark else ''),
        ))
    for code, (oee, target, remark) in mes.items():
        key = f'{code}|'
        if key in benchmarks:
            continue
        benchmarks[key] = asdict(BenchmarkThroughput(
            machine_code=code, material_number='', mes_oee=oee,
            note=f'MES-gemeten OEE {oee:.1%}'
                 + (f', doel {target:.0%}' if target else '')
                 + (f' — {remark}' if remark else ''),
        ))

    report.append('## 4. Benchmarks (bladen `MES_OEE Mills` en `PEER_Capacity`)\n')
    report.append(f'{len(benchmarks)} benchmarkregels: gemeten OEE per machine en PEER-doorzet per '
                  'machine × materiaal. Deze rekenen NIET mee — ze staan naast de norm in de werkbank, '
                  'zodat "actual vs norm" zichtbaar is.')
    if mes_skipped:
        report.append('')
        report.append('Overgeslagen werkcenters (geen eenduidige machine of geen meting): '
                      + ', '.join(sorted(set(mes_skipped))) + '.')
    report.append('')

    # 5. Normen-blad ---------------------------------------------------------
    normen = read_normen(wb[SHEET_NORMEN])
    report.append('## 5. Blad `Normen ` — bewust NIET in de database\n')
    report.append(f'{len(normen)} regels productie-unit × product met MES- en PEER-doorzet en de '
                  'OEE-opbouw (performance × availability). De productkolom bevat handelsnamen '
                  '("Ank125", "SG63"), geen materiaalnummers — koppelen aan de app-materialen kan '
                  'niet zonder de klant. Onderstaande tabel is de vraag, concreet gemaakt:')
    report.append('')
    report.append('| productie-unit | product | MES t/u | PEER t/u | perf. | avail. | OEE |')
    report.append('|---|---|---:|---:|---:|---:|---:|')
    for item in normen:
        def _f(v, pct=False):
            if v is None:
                return '—'
            return f'{v:.0%}' if pct else f'{v:g}'
        report.append(f'| {item["unit"]} | {item["product"]} | {_f(item["mes"])} | {_f(item["peer"])} '
                      f'| {_f(item["performance"], True)} | {_f(item["availability"], True)} '
                      f'| {_f(item["oee"], True)} |')
    report.append('')
    questions.append('Blad `Normen `: welk materiaalnummer hoort bij elke producthandelsnaam? '
                     'Daarna kunnen de MES-normen per machine × product de database in.')
    questions.append('Loonkosten per functiegroep staan niet in dit werkboek. Zonder tarief blijft '
                     'de kostenkolom van de werkbank leeg (er wordt niets verzonnen).')

    wb.close()

    datasets = {
        'fte_params': params,
        'fte_hours_per_year': effective_hours,
        'shift_hours': shift_hours,
        'staffing_norms': staffing_norms,
        'indirect_activities': activities,
        'benchmark_throughput': benchmarks,
        # Bewust leeg — zie het rapport.
        'labor_rates': {},
        'throughput_overrides': {},
        'machine_combinations': {},
    }
    return datasets, report, questions


# ── store bijwerken ─────────────────────────────────────────────────────────


def workbook_site(workbook: Path):
    """De site die het werkboek ZELF declareert, of None.

    Niet loader.config.site gebruiken: de loader vult daar een defaultsite in
    als de Config-sheet geen Site-regel heeft, en dan zou een werkboek van een
    andere site zonder Site-regel de sitecontrole passeren. De aanroeper
    behandelt None als weigering, niet als vrijbrief.
    """
    import pandas as pd

    try:
        config_sheet = pd.read_excel(str(workbook), sheet_name='Config')
    except Exception:
        return None
    for _, row in config_sheet.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        value = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
        if label == 'Site' and value is not None and str(value).strip():
            return str(value).strip()
    return None


def load_machines(workbook: Path):
    """Machines uit een MS_RECONC-werkboek, in store-formaat."""
    import contextlib
    import io

    from modules.data_loader import DataLoader

    with contextlib.redirect_stdout(io.StringIO()):
        loader = DataLoader(excel_file=str(workbook)).load_all()
    machines = []
    for machine in loader.machines.values():
        item = asdict(machine)
        item['shift_system'] = machine.shift_system.value
        machines.append(item)
    return machines, workbook_site(workbook)


def _merge_dataset(existing: dict, seeded: dict, overwrite: bool):
    """Voeg seedrecords toe zonder klantbewerkingen terug te draaien.

    Standaard worden alleen ONTBREKENDE sleutels toegevoegd. De hele opzet is
    dat de klant deze records daarna bijstelt (een activiteit aanzetten, een
    norm corrigeren); een tweede --apply die dat overschrijft maakt het script
    gevaarlijker dan nuttig. Met --overwrite-existing kan het wel, expliciet.
    """
    merged = dict(existing or {})
    stats = {'toegevoegd': [], 'ongewijzigd': [], 'overschreven': []}
    for key, value in (seeded or {}).items():
        if key not in merged:
            merged[key] = value
            stats['toegevoegd'].append(key)
        elif merged[key] == value:
            stats['ongewijzigd'].append(key)
        elif overwrite:
            merged[key] = value
            stats['overschreven'].append(key)
        else:
            stats['ongewijzigd'].append(key)
    return merged, stats


def _describe(name: str, stats: dict, overwrite: bool) -> str:
    parts = [f'{len(stats["toegevoegd"])} toegevoegd',
             f'{len(stats["ongewijzigd"])} ongemoeid gelaten']
    if stats['overschreven']:
        shown = ', '.join(stats['overschreven'][:6])
        more = '...' if len(stats['overschreven']) > 6 else ''
        parts.append(f'{len(stats["overschreven"])} OVERSCHREVEN ({shown}{more})')
    line = f'{name}: ' + ', '.join(parts)
    if stats['ongewijzigd'] and not overwrite:
        line += (' - bestaande records blijven staan '
                 '(--overwrite-existing vervangt ze)')
    return line


def _merge_machines(existing: list, seeded: list, overwrite: bool):
    """Machines per VELD samenvoegen in plaats van integraal vervangen.

    Een in de app gecorrigeerde OEE of naam mag niet terugspringen naar de
    werkboekwaarde: OEE schaalt de capaciteit rechtstreeks, dus dat verschuift
    Line 09-12 en de hele werkbank zonder melding.
    """
    by_code = {m['machine_code']: dict(m) for m in (existing or [])}
    notes = []
    for machine in seeded or []:
        code = machine['machine_code']
        current = by_code.get(code)
        if current is None:
            by_code[code] = dict(machine)
            notes.append(f'+{code}')
            continue
        incoming = dict(machine)
        # Beschikbaarheid is MAANDdata; die van de store wint, tenzij leeg.
        # Is de store leeg en het werkboek gevuld, dan is het geen conflict
        # maar een AANVULLING — die moet landen, ook zonder --overwrite, en
        # hij moet in het rapport staan.
        fills_month_data = (not current.get('availability_by_period')
                            and bool(incoming.get('availability_by_period')))
        if current.get('availability_by_period'):
            incoming['availability_by_period'] = current['availability_by_period']
        changed = [f for f, v in incoming.items()
                   if f != 'availability_by_period' and current.get(f) != v]
        if fills_month_data:
            current['availability_by_period'] = incoming['availability_by_period']
            notes.append(f'~{code} (beschikbaarheid aangevuld)')
        if not changed:
            continue
        if overwrite:
            current.update(incoming)
            notes.append(f'~{code} ({", ".join(changed)})')
        else:
            notes.append(f'={code} ongewijzigd gelaten (werkboek wijkt af op '
                         f'{", ".join(changed)})')
    return [by_code[c] for c in sorted(by_code)], notes


def apply_seed(store_path: Path, datasets: dict, machines, verify_only: bool,
               keep_fte_hours: bool = False, overwrite: bool = False,
               expected_site: str = ''):
    """Voeg de seed samen met de bestaande store. Retourneert (master, wijzigingen)."""
    from ui import master_store

    master_store.set_store_path(store_path)
    record = master_store.get_current_master_record()
    if record is None:
        raise SystemExit(f'Geen masterdata-store gevonden op {store_path}. '
                         'Importeer eerst masterdata in de app.')
    master = json.loads(json.dumps(record['master']))
    changes = []

    # Sitecontrole: met SOP_APP_DATA_DIR gezet wijst de default store naar een
    # ZUSTERSITE. Maastricht-normen daar inschrijven is stille datavervuiling.
    # Fail-closed: ook een store ZONDER site wordt geweigerd — "we weten niet
    # welke site dit is" is geen reden om er dan maar in te schrijven (een
    # legacy of handbewerkte store mist het config-blok juist het vaakst).
    store_site = str((master.get('config') or {}).get('site') or '').strip()
    if expected_site:
        if not store_site:
            raise SystemExit(
                f'Doelstore op {store_path} heeft geen site in zijn config; '
                f'niet vast te stellen of dit de {expected_site}-store is. '
                'Gebruik --force-site als dit toch de bedoeling is.')
        if store_site != expected_site:
            raise SystemExit(
                f'Doelstore is site {store_site}, maar deze seed hoort bij '
                f'{expected_site} ({store_path}). Gebruik --force-site als dit '
                f'toch de bedoeling is.')

    fte = dict(master.get('fte') or {})
    old_hours = float(fte.get('fte_hours_per_year') or 0.0)
    new_hours = datasets['fte_hours_per_year']
    existing_params = fte.get('params') or {}

    # Gate op "heeft de KLANT hier iets aan gedaan", niet op "staat er iets".
    # serialize_master schrijft ALTIJD een params-blok, gevuld met defaults,
    # dus 'er staan al parameters' was nooit onwaar en de allereerste --apply
    # sloeg juist het getal over waar de hele seed om begon.
    from modules.models import FTE_PARAM_DEFAULTS

    edited_params = {name: value for name, value in existing_params.items()
                     if name in FTE_PARAM_DEFAULTS
                     and abs(float(value or 0.0)
                             - float(FTE_PARAM_DEFAULTS[name] or 0.0)) > 1e-12}
    if edited_params and not overwrite:
        changes.append(
            'fte.params ONGEWIJZIGD — de klant heeft '
            + ', '.join(sorted(edited_params)) + ' aangepast '
            '(--overwrite-existing vervangt ze alsnog)')
    else:
        old_rate = existing_params.get('utilization_rate')
        new_rate = datasets['fte_params'].get('utilization_rate')
        fte['params'] = datasets['fte_params']
        changes.append(f'fte.params gezet ({len(datasets["fte_params"])} parameters); '
                       'de bruto->netto-velden documenteren alleen de afleiding')
        # utilization_rate zit in datzelfde blok maar REKENT mee: FteResult
        # deelt de benodigde FTE erdoor voor 'bij bezettingsdoel'. Die mag niet
        # onder 'verandert geen getal' verdwijnen.
        if old_rate is not None and abs(float(old_rate) - float(new_rate or 0.0)) > 1e-12:
            changes.append(
                f'**bezettingsgraad {float(old_rate):g} -> {float(new_rate):g}** — '
                f'die rekent WEL mee: de benodigde FTE bij bezettingsdoel schuift '
                f'met {abs(1/float(new_rate) - 1/float(old_rate)) * 100 / (1/float(old_rate)):.1f}%.')

    shift_added, shift_replaced = [], []
    for name, hours in datasets['shift_hours'].items():
        current = (fte.get('shift_hours') or {}).get(name)
        if current is None:
            fte.setdefault('shift_hours', {})[name] = hours
            shift_added.append(name)
        elif overwrite and abs(float(current) - float(hours)) > 1e-9:
            fte.setdefault('shift_hours', {})[name] = hours
            shift_replaced.append(f'{name} {float(current):g} -> {float(hours):g}')
    # Ploeguren schalen de beschikbare uren van ELKE machine in dat systeem;
    # ze stil muteren is precies wat het wijzigingsoverzicht moet voorkomen.
    if shift_added or shift_replaced:
        parts = []
        if shift_added:
            parts.append(f'{len(shift_added)} toegevoegd ({", ".join(shift_added)})')
        if shift_replaced:
            parts.append('OVERSCHREVEN: ' + '; '.join(shift_replaced)
                         + ' — dit schaalt de beschikbare uren van elke machine '
                           'in dat ploegensysteem')
        changes.append('ploeguren: ' + ', '.join(parts))
    else:
        changes.append('ploeguren ONGEWIJZIGD')

    if keep_fte_hours or abs(old_hours - new_hours) < 0.005:
        changes.append(f'uren per FTE ONGEWIJZIGD op {old_hours:g}')
    elif edited_params and not overwrite:
        changes.append(f'uren per FTE ONGEWIJZIGD op {old_hours:g} (werkboek zegt '
                       f'{new_hours:.2f}; --overwrite-existing past het aan)')
    else:
        fte['fte_hours_per_year'] = new_hours
        shift = (f' Dat schuift ELK FTE-getal met '
                 f'{abs(new_hours - old_hours) / old_hours:.2%}.' if old_hours > 0
                 else ' De store had er nog geen.')
        changes.append(
            f'**uren per FTE {old_hours:g} -> {new_hours:.2f}** - de exacte waarde uit '
            f'het werkboek in plaats van de afgeronde.{shift} Draai met '
            f'--keep-current-fte-hours om de huidige waarde te behouden.')
    master['fte'] = fte

    for name in ('staffing_norms', 'indirect_activities', 'benchmark_throughput'):
        merged, stats = _merge_dataset(master.get(name), datasets[name], overwrite)
        master[name] = merged
        changes.append(_describe(name, stats, overwrite))

    for name in ('labor_rates', 'throughput_overrides', 'machine_combinations'):
        master.setdefault(name, {})

    # Groepsmaterialen een leesbare naam geven, alleen waar er nog geen staat.
    renamed = []
    for material in master.get('materials') or []:
        target = GROUP_NAME.get(str(material.get('material_number')))
        current = str(material.get('name') or '').strip()
        if target and (not current or current.lower() == 'nan'):
            material['name'] = target
            renamed.append(f'{material["material_number"]} -> "{target}"')
    if renamed:
        changes.append(f'machinegroepen een naam gegeven ({len(renamed)}): '
                       + ', '.join(renamed))

    if machines is not None:
        before = len(master.get('machines') or [])
        master['machines'], notes = _merge_machines(master.get('machines'), machines,
                                                    overwrite)
        changes.append(f'machines: {before} -> {len(master["machines"])}'
                       + (f' | {"; ".join(notes)}' if notes else ''))

    if verify_only:
        return master, changes

    from modules.master_data import hydrate_loader

    class _Probe:
        pass

    hydrate_loader(_Probe(), master)   # validatie door hydratie, als de PATCH-route

    # Compare-and-swap op de storeversie. LET OP de reikwijdte: de lock
    # hieronder is een threading.Lock die per PROCES bestaat, dus tussen de
    # seed en een draaiende app beschermt hij niets — hij houdt alleen deze
    # run intern consistent. Wat wél helpt is de versiecheck: die vangt een
    # wijziging die vóór dit punt is gedaan. Er blijft een klein venster
    # tussen de check en de save; draai de seed daarom bij voorkeur met de
    # app dicht.
    from ui.routes.master_data import _master_mutation_lock

    with _master_mutation_lock:
        current = master_store.get_current_master_record()
        if (current or {}).get('version') != record.get('version'):
            raise SystemExit(
                f'De masterdata is tijdens deze seed gewijzigd (versie '
                f'{record.get("version")} -> {(current or {}).get("version")}). '
                f'Er is NIETS geschreven. Draai opnieuw.')
        backup = store_path.with_name(
            f'{store_path.name}.backup-{datetime.now().strftime("%Y%m%d%H%M%S")}')
        shutil.copy2(store_path, backup)
        master_store.save_master_store(store_path, master, previous=current, edited=True)
    changes.append(f'back-up geschreven naar {backup.name}')

    # De spiegel op schijf is het referentiedocument; die moet meteen de
    # nieuwe stand tonen.
    from ui import master_mirror

    master_mirror.set_mirror_dir(store_path.parent)
    status = master_mirror.refresh_mirror()
    if status.get('stale'):
        changes.append(f'LET OP: masterwerkboek niet bijgewerkt - {status.get("reason")}')
    else:
        changes.append(f'masterwerkboek ververst: {Path(status["path"]).name}')
    return master, changes


def main() -> int:
    # Windows-consoles draaien standaard cp1252. Zonder deze regel eindigde
    # het script met een UnicodeEncodeError op het eerste niet-ASCII teken —
    # NA de storemutatie, dus met exit 1, een gewijzigde database en een
    # onzichtbaar wijzigingsrapport. De redelijke reactie daarop ("mislukt,
    # nog eens draaien") was precies de gevaarlijkste.
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding='utf-8', errors='replace')
        except (AttributeError, ValueError):
            pass

    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--workbook', type=Path, default=DEFAULT_WORKBOOK,
                        help='het OEE/FTE-klantmodel (.xlsx)')
    parser.add_argument('--store', type=Path,
                        help='pad naar master_store.json (default: %LOCALAPPDATA%\\SOPPlanningEngine)')
    parser.add_argument('--machines-workbook', type=Path,
                        help='MS_RECONC-werkboek om de machinelijst uit te vullen')
    parser.add_argument('--keep-current-fte-hours', action='store_true',
                        help='laat fte_hours_per_year staan zoals hij is')
    parser.add_argument('--overwrite-existing', action='store_true',
                        help='vervang records die al in de store staan. ZONDER deze vlag '
                             'voegt de seed alleen ontbrekende records toe en blijven '
                             'klantbewerkingen (aangezette activiteiten, gecorrigeerde '
                             'normen, aangepaste machine-OEE) staan.')
    parser.add_argument('--force-site', action='store_true',
                        help='sla de sitecontrole over (normaal weigert de seed een store '
                             'van een andere site dan ' + EXPECTED_SITE + ')')
    parser.add_argument('--apply', action='store_true',
                        help='schrijf naar de store (zonder deze vlag verandert er niets)')
    parser.add_argument('--report', type=Path,
                        help='schrijf het rapport naar dit bestand. Het rapport bevat '
                             'klantcijfers (volumes, normen, bezetting) en mag daarom '
                             'NIET in de repo staan; een pad binnen de repo wordt geweigerd.')
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f'Werkboek niet gevonden: {args.workbook}')
        return 1

    store_path = args.store
    if store_path is None:
        from ui.paths import default_app_data_root
        store_path = default_app_data_root() / 'master_store.json'

    if args.report:
        repo_root = Path(__file__).resolve().parent.parent
        try:
            args.report.resolve().relative_to(repo_root)
        except ValueError:
            pass
        else:
            print(f'Weigering: {args.report} ligt in de repo. Het rapport bevat '
                  f'klantcijfers. Kies een pad daarbuiten, bijvoorbeeld naast de store:'
                  f'\n  --report "{store_path.parent / "seed-f2-cf-rapport.md"}"')
            return 1

    machines = None
    site = None
    if args.machines_workbook:
        machines, site = load_machines(args.machines_workbook)
        # Fail-closed: 'site is None' betekent dat de Config-sheet geen
        # Site-regel heeft. Dat was eerder een vrijbrief (de loader-default
        # maakte er stil NLX1 van), terwijl juist een vreemd of verminkt
        # werkboek die regel mist.
        if not args.force_site and site != EXPECTED_SITE:
            beschrijving = (f'is van site {site}' if site
                            else 'heeft geen Site-regel in de Config-sheet')
            print(f'Weigering: {args.machines_workbook.name} {beschrijving}, '
                  f'maar deze seed hoort bij {EXPECTED_SITE}. Gebruik --force-site '
                  f'als dit toch de bedoeling is.')
            return 1

    # Machines om de vertaaltabel tegen te controleren: die uit het opgegeven
    # werkboek, anders die al in de store staan.
    sap_machines = {m['machine_code']: m for m in (machines or [])}
    if not sap_machines:
        from ui import master_store
        master_store.set_store_path(store_path)
        record = master_store.get_current_master_record()
        sap_machines = {m['machine_code']: m
                        for m in ((record or {}).get('master', {}).get('machines') or [])}

    try:
        datasets, report, questions = build_seed(args.workbook, sap_machines)
    except SeedError as exc:
        print(f'Seed afgebroken: {exc}')
        return 1

    if machines is not None:
        report.insert(0, f'Machinelijst uit `{args.machines_workbook.name}` (site {site}): '
                         f'{len(machines)} machines.\n')
    elif not sap_machines:
        report.insert(0, '**Let op: de masterdata bevat GEEN machines.** Zonder machines '
                         'rekent de app geen capaciteit en dus ook geen FTE. Draai met '
                         '`--machines-workbook <MS_RECONC.xlsm>` om ze te vullen.\n')

    header = [
        f'# Seed F2-CF masterdata - {EXPECTED_SITE}',
        '',
        f'Bron: `{args.workbook.name}`',
        f'Store: `{store_path}`',
        f'Modus: {"TOEPASSEN" if args.apply else "droogdraaien (er verandert niets)"}'
        + (' + OVERSCHRIJVEN van bestaande records' if args.overwrite_existing else ''),
        f'Datum: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
        '',
    ]

    try:
        _, changes = apply_seed(store_path, datasets, machines,
                                verify_only=not args.apply,
                                keep_fte_hours=args.keep_current_fte_hours,
                                overwrite=args.overwrite_existing,
                                expected_site='' if args.force_site else EXPECTED_SITE)
    except SystemExit as exc:
        print(str(exc))
        return 1

    tail = ['## 6. Wat er met de database gebeurt', '']
    tail += [f'- {c}' for c in changes]
    tail += ['', '## 7. Openstaande vragen aan de klant', '']
    tail += [f'{i}. {q}' for i, q in enumerate(dict.fromkeys(questions), 1)]
    tail += ['']

    text = '\n'.join(header + report + tail)
    if args.report:
        # Schrijffouten mogen nooit ná een geslaagde mutatie alsnog exit 1
        # geven: de operator concludeert dan 'mislukt' en draait opnieuw.
        try:
            args.report.parent.mkdir(parents=True, exist_ok=True)
            args.report.write_text(text, encoding='utf-8')
            print(f'Rapport geschreven naar {args.report}')
        except OSError as exc:
            print(f'LET OP: rapport kon niet worden geschreven ({exc}). '
                  f'De database is wel bijgewerkt; het rapport volgt hieronder.')
    # Nooit exit 1 op een presentatiefout nadat de database al gemuteerd is.
    try:
        print(text)
    except Exception as exc:  # pragma: no cover - afhankelijk van de console
        print(f'(rapport kon niet worden getoond: {exc}; '
              f'gebruik --report om het naar een bestand te schrijven)')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
